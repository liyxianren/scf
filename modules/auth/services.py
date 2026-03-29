"""自动排课、删除清理和账号初始化服务。"""
import io
import json
import re
import secrets
import unicodedata
from datetime import datetime, date, timedelta, time, timezone
from itertools import combinations, product
from math import ceil
from zoneinfo import ZoneInfo

from sqlalchemy import and_, or_

from extensions import db
from modules.auth.availability_ai_services import (
    build_availability_intake_summary,
    parse_availability_intake,
    resolve_availability_evidence_items,
)


FEEDBACK_PREFIX = "[排课反馈]"
SCHEDULE_PREFIX = "[排课方案]"
BUSINESS_TIMEZONE = ZoneInfo("Asia/Shanghai")
UTC = timezone.utc
LEGACY_ENROLLMENT_NOTE_PATTERN = re.compile(r'(?<!\d)报名#\s*(\d+)(?!\d)')
IMPORTED_SCHEDULE_FEEDBACK_START_DATE = date(2026, 4, 1)
DELIVERY_URGENCY_LABELS = {
    'normal': '常规',
    'rush': '冲刺',
}
TEACHER_WORK_MODE_PART_TIME = 'part_time'
TEACHER_WORK_MODE_FULL_TIME = 'full_time'
TEACHER_WORK_MODE_LABELS = {
    TEACHER_WORK_MODE_PART_TIME: '兼职老师',
    TEACHER_WORK_MODE_FULL_TIME: '全职老师',
}
DEFAULT_FULL_TIME_WORKING_TEMPLATE = [
    {'day': 2, 'start': '10:00', 'end': '18:00'},
    {'day': 3, 'start': '10:00', 'end': '18:00'},
    {'day': 4, 'start': '10:00', 'end': '18:00'},
    {'day': 5, 'start': '10:00', 'end': '18:00'},
    {'day': 6, 'start': '10:00', 'end': '18:00'},
]


def get_business_now():
    """返回系统当前使用的本地业务时间。"""
    return datetime.now(BUSINESS_TIMEZONE).replace(tzinfo=None)


def serialize_business_datetime(value, *, assume_utc=True):
    """把数据库中的时间值格式化成带业务时区偏移的 ISO 字符串。"""
    if value is None:
        return None

    if value.tzinfo is None:
        source_tz = UTC if assume_utc else BUSINESS_TIMEZONE
        aware_value = value.replace(tzinfo=source_tz)
    else:
        aware_value = value

    return aware_value.astimezone(BUSINESS_TIMEZONE).isoformat()


def get_business_today():
    return get_business_now().date()


def schedule_requires_course_feedback(schedule):
    if not schedule or not schedule.date:
        return False
    if getattr(schedule, 'is_cancelled', False):
        return False
    if getattr(schedule, 'import_run_id', None) and schedule.date < IMPORTED_SCHEDULE_FEEDBACK_START_DATE:
        return False
    return True


def get_course_feedback_skip_reason(schedule):
    if not schedule or not schedule.date:
        return '课次信息不完整，暂不生成课程反馈待办'
    if getattr(schedule, 'is_cancelled', False):
        return (getattr(schedule, 'cancel_reason', None) or '课次已取消，当前无需提交课程反馈').strip()
    if getattr(schedule, 'import_run_id', None) and schedule.date < IMPORTED_SCHEDULE_FEEDBACK_START_DATE:
        return (
            f'{IMPORTED_SCHEDULE_FEEDBACK_START_DATE.isoformat()} 前导入的历史课次'
            '不要求补录课程反馈'
        )
    return None


def generate_intake_token():
    """生成学生填表链接的 token。"""
    return secrets.token_urlsafe(32)


def _coerce_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {'1', 'true', 'yes', 'y', 'on'}:
            return True
        if normalized in {'0', 'false', 'no', 'n', 'off'}:
            return False
    return bool(value)


def _serialize_json_field(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def _normalize_teacher_work_mode(value, *, default=TEACHER_WORK_MODE_PART_TIME):
    normalized = str(value or '').strip().lower()
    if not normalized:
        return default
    alias_map = {
        'part_time': TEACHER_WORK_MODE_PART_TIME,
        'part-time': TEACHER_WORK_MODE_PART_TIME,
        '兼职': TEACHER_WORK_MODE_PART_TIME,
        'parttime': TEACHER_WORK_MODE_PART_TIME,
        'full_time': TEACHER_WORK_MODE_FULL_TIME,
        'full-time': TEACHER_WORK_MODE_FULL_TIME,
        'fulltime': TEACHER_WORK_MODE_FULL_TIME,
        '全职': TEACHER_WORK_MODE_FULL_TIME,
    }
    result = alias_map.get(normalized)
    if not result:
        raise ValueError('老师工作模式必须是 full_time 或 part_time')
    return result


def teacher_work_mode_label(value):
    try:
        normalized = _normalize_teacher_work_mode(value)
    except ValueError:
        normalized = TEACHER_WORK_MODE_PART_TIME
    return TEACHER_WORK_MODE_LABELS.get(normalized, TEACHER_WORK_MODE_LABELS[TEACHER_WORK_MODE_PART_TIME])


def _normalize_default_working_template(value):
    if value in (None, '', []):
        return [], []
    normalized, errors = _validate_available_slot_entries(value)
    return normalized, [
        str(item).replace('可补课时段', '默认工作时段')
        for item in (errors or [])
    ]


def get_company_full_time_working_template():
    return [dict(item) for item in DEFAULT_FULL_TIME_WORKING_TEMPLATE]


def _resolve_teacher_user_record(teacher_or_user):
    from modules.auth.models import User

    if teacher_or_user is None:
        return None
    if isinstance(teacher_or_user, User):
        return teacher_or_user
    if hasattr(teacher_or_user, 'teacher') and getattr(teacher_or_user, 'teacher', None):
        return teacher_or_user.teacher
    if hasattr(teacher_or_user, 'teacher_id') and getattr(teacher_or_user, 'teacher_id', None):
        return db.session.get(User, teacher_or_user.teacher_id)
    if hasattr(teacher_or_user, 'id') and hasattr(teacher_or_user, 'role'):
        return teacher_or_user
    return db.session.get(User, teacher_or_user)


def resolve_teacher_work_mode(teacher_or_user):
    teacher = _resolve_teacher_user_record(teacher_or_user)
    if not teacher:
        return TEACHER_WORK_MODE_PART_TIME
    return _normalize_teacher_work_mode(
        getattr(teacher, 'teacher_work_mode', None),
        default=TEACHER_WORK_MODE_PART_TIME,
    )


def resolve_teacher_default_working_template(teacher_or_user):
    teacher = _resolve_teacher_user_record(teacher_or_user)
    mode = resolve_teacher_work_mode(teacher)
    if mode != TEACHER_WORK_MODE_FULL_TIME:
        return []

    raw_value = getattr(teacher, 'default_working_template_json', None) if teacher else None
    template, errors = _normalize_default_working_template(raw_value)
    if template and not errors:
        return template
    return get_company_full_time_working_template()


def teacher_availability_ready(teacher_or_user):
    teacher = _resolve_teacher_user_record(teacher_or_user)
    if not teacher:
        return False
    if resolve_teacher_work_mode(teacher) == TEACHER_WORK_MODE_FULL_TIME:
        return True
    return bool(_load_teacher_available_ranges(getattr(teacher, 'id', None)))


def teacher_requires_manual_proposal(teacher_or_user):
    return resolve_teacher_work_mode(teacher_or_user) != TEACHER_WORK_MODE_FULL_TIME


def teacher_auto_accept_enabled(teacher_or_user):
    return resolve_teacher_work_mode(teacher_or_user) == TEACHER_WORK_MODE_FULL_TIME


def enrollment_requires_teacher_confirmation(enrollment):
    if not enrollment:
        return False
    if teacher_requires_manual_proposal(getattr(enrollment, 'teacher', None) or enrollment.teacher_id):
        return True
    risk_assessment = _enrollment_json_field(getattr(enrollment, 'risk_assessment', None)) or {}
    return bool(risk_assessment.get('teacher_confirmation_required'))


def _teacher_work_context(teacher_or_user):
    teacher = _resolve_teacher_user_record(teacher_or_user)
    mode = resolve_teacher_work_mode(teacher)
    default_working_template = resolve_teacher_default_working_template(teacher)
    using_company_template = bool(
        mode == TEACHER_WORK_MODE_FULL_TIME
        and not getattr(teacher, 'default_working_template_json', None)
    )
    return {
        'teacher_work_mode': mode,
        'teacher_work_mode_label': teacher_work_mode_label(mode),
        'default_working_template': default_working_template,
        'default_working_template_summary': _summarize_available_slots(default_working_template),
        'using_company_template': using_company_template,
        'availability_source': (
            'company_template'
            if mode == TEACHER_WORK_MODE_FULL_TIME and using_company_template
            else ('teacher_template' if mode == TEACHER_WORK_MODE_FULL_TIME else 'manual_availability')
        ),
        'availability_ready': teacher_availability_ready(teacher),
        'teacher_auto_accept_enabled': teacher_auto_accept_enabled(teacher),
        'requires_teacher_proposal': teacher_requires_manual_proposal(teacher),
    }


def _normalize_delivery_preference(value, *, required=False):
    from modules.oa.services import normalize_delivery_mode

    raw_value = value
    if raw_value in (None, '') and required:
        raise ValueError('请先选择线上或线下上课方式')
    if raw_value in (None, '') and not required:
        return normalize_delivery_mode(
            raw_value,
            allow_unknown=True,
            default='unknown',
        )
    try:
        return normalize_delivery_mode(
            raw_value,
            allow_unknown=False,
            default='unknown',
        )
    except ValueError:
        if raw_value in (None, ''):
            raise ValueError('请先选择线上或线下上课方式')
        raise ValueError('上课方式必须是 online 或 offline')


def _normalize_delivery_urgency(value, *, required=False):
    normalized = str(value or '').strip().lower()
    if not normalized:
        if required:
            raise ValueError('请先选择交付节奏')
        return 'normal'
    alias_map = {
        'normal': 'normal',
        'regular': 'normal',
        '常规': 'normal',
        'rush': 'rush',
        'urgent': 'rush',
        '冲刺': 'rush',
        '比赛': 'rush',
    }
    result = alias_map.get(normalized)
    if not result:
        raise ValueError('交付节奏必须是 normal 或 rush')
    return result


def _parse_optional_date(value, *, field_label='日期'):
    if value in (None, '', 'null'):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError(f'{field_label}格式错误，请使用 YYYY-MM-DD') from exc


def _parse_positive_int(value, *, default=1, minimum=1):
    if value in (None, ''):
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError('每周节次数必须是整数') from exc
    return max(parsed, minimum)


def _enrollment_json_field(value):
    if value in (None, ''):
        return None
    if isinstance(value, str):
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return None
    return value


def _resolve_teacher_user(teacher_id=None, teacher_name=None):
    from modules.auth.models import User
    from modules.oa.services import resolve_schedule_teacher_reference

    if teacher_id:
        teacher = db.session.get(User, teacher_id)
        if not teacher:
            return None, '授课老师不存在'
        return teacher, None

    if teacher_name:
        teacher, _, _, error = resolve_schedule_teacher_reference(teacher_name)
        if error or not teacher:
            return None, error or f'未找到老师: {teacher_name}'
        return teacher, None

    return None, '缺少 teacher_id 或 teacher_name'


def _teacher_identity_names(user):
    if not user:
        return []
    values = []
    for raw in [getattr(user, 'display_name', None), getattr(user, 'username', None)]:
        text = str(raw or '').strip()
        if text:
            values.append(text)
    return list(dict.fromkeys(values))


def _teacher_schedule_identity_filter(schedule_model, user):
    if not user or not getattr(user, 'id', None):
        return schedule_model.id == None
    clauses = [schedule_model.teacher_id == user.id]
    names = _teacher_identity_names(user)
    if names:
        clauses.append(and_(schedule_model.teacher_id.is_(None), schedule_model.teacher.in_(names)))
    return or_(*clauses)


def _schedule_matches_teacher_actor(actor, schedule=None, *, teacher_id=None, teacher_name=None):
    if not (
        actor
        and getattr(actor, 'is_authenticated', False)
        and getattr(actor, 'role', None) in {'teacher', 'admin'}
    ):
        return False
    target_teacher_id = teacher_id if teacher_id is not None else getattr(schedule, 'teacher_id', None)
    if target_teacher_id is not None:
        return actor.id == target_teacher_id
    target_teacher_name = str(
        teacher_name if teacher_name is not None else getattr(schedule, 'teacher', None) or ''
    ).strip()
    return bool(target_teacher_name and target_teacher_name in _teacher_identity_names(actor))


def _hydrate_schedule_teacher_id(schedule):
    if not schedule or getattr(schedule, 'teacher_id', None):
        return getattr(schedule, 'teacher_id', None)
    teacher_name = str(getattr(schedule, 'teacher', None) or '').strip()
    if not teacher_name:
        return None
    teacher, error = _resolve_teacher_user(teacher_name=teacher_name)
    if teacher and not error:
        schedule.teacher_id = teacher.id
        return teacher.id
    return None


def student_schedule_profile_clause(profile_ids, *, schedule_model=None):
    from modules.auth.models import Enrollment
    from modules.oa.models import CourseSchedule

    schedule_model = schedule_model or CourseSchedule
    if isinstance(profile_ids, (list, tuple, set)):
        normalized_profile_ids = [item for item in profile_ids if item]
    elif profile_ids:
        normalized_profile_ids = [profile_ids]
    else:
        normalized_profile_ids = []

    if not normalized_profile_ids:
        return schedule_model.id.is_(None)

    return or_(
        schedule_model.student_profile_id_snapshot.in_(normalized_profile_ids),
        and_(
            schedule_model.student_profile_id_snapshot.is_(None),
            schedule_model.enrollment.has(Enrollment.student_profile_id.in_(normalized_profile_ids)),
        ),
    )


def _resolve_schedule_student_profile_id(schedule, *, enrollment=None):
    from modules.auth.models import Enrollment

    linked_enrollment = enrollment
    if linked_enrollment is None and schedule and getattr(schedule, 'enrollment_id', None):
        linked_enrollment = getattr(schedule, 'enrollment', None) or db.session.get(Enrollment, schedule.enrollment_id)

    if linked_enrollment and linked_enrollment.student_profile_id:
        return linked_enrollment.student_profile_id

    latest_leave = _latest_leave_request(schedule)
    if latest_leave and latest_leave.enrollment and latest_leave.enrollment.student_profile_id:
        return latest_leave.enrollment.student_profile_id

    return None


def sync_schedule_student_snapshot(schedule, *, enrollment=None, preserve_history=None, force=False):
    if not schedule:
        return None

    if preserve_history is None:
        preserve_history = schedule_has_historical_facts(schedule)

    next_profile_id = _resolve_schedule_student_profile_id(schedule, enrollment=enrollment)
    current_snapshot = getattr(schedule, 'student_profile_id_snapshot', None)
    if preserve_history and current_snapshot is not None and not force:
        return current_snapshot

    schedule.student_profile_id_snapshot = next_profile_id
    return next_profile_id


def _schedule_effective_student_profile_id(schedule):
    if not schedule:
        return None
    snapshot_profile_id = getattr(schedule, 'student_profile_id_snapshot', None)
    if snapshot_profile_id is not None:
        return snapshot_profile_id
    enrollment = getattr(schedule, 'enrollment', None)
    if enrollment and enrollment.student_profile_id:
        return enrollment.student_profile_id
    return None


def create_enrollment_record(data):
    """创建报名记录并生成 intake 链接。"""
    from modules.auth.models import Enrollment

    student_name = (data.get('student_name') or '').strip()
    course_name = (data.get('course_name') or '').strip()
    teacher_name = (data.get('teacher_name') or data.get('teacher') or '').strip()
    teacher_id = data.get('teacher_id')

    if not student_name:
        return None, None, '缺少必填字段: student_name'
    if not course_name:
        return None, None, '缺少必填字段: course_name'

    teacher, error = _resolve_teacher_user(teacher_id=teacher_id, teacher_name=teacher_name)
    if error:
        return None, None, error

    token_days = data.get('token_ttl_days', 7)
    try:
        token_days = int(token_days)
    except (TypeError, ValueError):
        token_days = 7

    token = generate_intake_token()
    try:
        delivery_preference = _normalize_delivery_preference(
            data.get('delivery_preference'),
            required=False,
        )
        delivery_urgency = _normalize_delivery_urgency(data.get('delivery_urgency'))
        sessions_per_week = _parse_positive_int(data.get('sessions_per_week'), default=1, minimum=1)
        target_finish_date = _parse_optional_date(
            data.get('target_finish_date'),
            field_label='目标完成日',
        )
    except ValueError as exc:
        return None, None, str(exc)
    if delivery_urgency == 'rush' and not target_finish_date:
        return None, None, '冲刺交付必须填写目标完成日'
    enrollment = Enrollment(
        student_name=student_name,
        course_name=course_name,
        teacher_id=teacher.id,
        delivery_preference=delivery_preference,
        delivery_urgency=delivery_urgency,
        target_finish_date=target_finish_date,
        total_hours=data.get('total_hours'),
        hours_per_session=data.get('hours_per_session', 2.0),
        sessions_per_week=sessions_per_week,
        notes=data.get('notes'),
        intake_token=token,
        token_expires_at=get_business_now() + timedelta(days=max(token_days, 1)),
        status='pending_info',
    )
    db.session.add(enrollment)
    db.session.commit()
    return enrollment, f'/auth/intake/{token}', None


def _normalize_username_base(student_name):
    normalized = unicodedata.normalize('NFKC', str(student_name or '').strip())
    normalized = re.sub(r'\s+', '', normalized)
    normalized = ''.join(
        char for char in normalized
        if char.isalnum() or char in {'-', '_', '.'}
    )
    return normalized[:100] or 'student'


def _build_student_username(student_name):
    from modules.auth.models import User

    base_username = _normalize_username_base(student_name)
    username = base_username
    suffix = 2
    while User.query.filter_by(username=username).first():
        suffix_text = str(suffix)
        username = f'{base_username[:100-len(suffix_text)]}{suffix_text}'
        suffix += 1
    return username


def _sync_student_user(user, student_name, phone):
    if not user:
        return
    user.display_name = student_name
    if phone:
        user.phone = phone


def _resolve_or_create_student_account(enrollment, student_name, phone):
    from modules.auth.models import User

    if not enrollment:
        return None, None

    linked_profile = enrollment.student_profile
    linked_user = linked_profile.user if linked_profile and linked_profile.user_id else None

    if linked_user:
        _sync_student_user(linked_user, student_name, phone)
        account_info = {
            'username': linked_user.username,
            'password': None,
            'user_id': linked_user.id,
        }
        return linked_user, account_info

    username = _build_student_username(student_name)
    student_user = User(
        username=username,
        display_name=student_name,
        role='student',
        phone=phone,
    )
    student_user.set_password('scf123')
    db.session.add(student_user)
    db.session.flush()
    if linked_profile and linked_profile.user_id is None:
        linked_profile.user_id = student_user.id
    account_info = {
        'username': username,
        'password': 'scf123',
        'user_id': student_user.id,
    }
    return student_user, account_info


def _apply_student_profile_fields(profile, data, *, preserve_missing=False):
    slots = data.get('available_slots')
    if slots is None and 'available_times' in data:
        slots = data.get('available_times')

    excluded_dates = data.get('excluded_dates')
    if 'name' in data or not preserve_missing:
        profile.name = data.get('name', profile.name)
    if 'grade' in data or not preserve_missing:
        profile.grade = data.get('grade', profile.grade if preserve_missing else None)
    if 'school' in data or not preserve_missing:
        profile.school = data.get('school', profile.school if preserve_missing else None)
    if 'phone' in data or not preserve_missing:
        profile.phone = data.get('phone', profile.phone)
    if 'parent_phone' in data or not preserve_missing:
        profile.parent_phone = data.get('parent_phone', profile.parent_phone if preserve_missing else None)
    if 'has_experience' in data or not preserve_missing:
        profile.has_experience = _coerce_bool(
            data.get('has_experience'),
            default=profile.has_experience if preserve_missing else False,
        )
    if 'experience_detail' in data or not preserve_missing:
        profile.experience_detail = data.get(
            'experience_detail',
            profile.experience_detail if preserve_missing else None,
        )
    if slots is not None:
        profile.available_slots = _serialize_json_field(slots)
    elif not preserve_missing:
        profile.available_slots = None
    if excluded_dates is not None:
        serialized = _serialize_json_field(excluded_dates)
        profile.excluded_dates = None if serialized in (None, '[]', '') else serialized
    elif not preserve_missing:
        profile.excluded_dates = None
    if 'notes' in data or not preserve_missing:
        profile.notes = data.get('notes', profile.notes if preserve_missing else None)


def submit_enrollment_intake(enrollment, data):
    """提交报名对应的学生信息表。"""
    from modules.auth.models import StudentProfile

    if not enrollment:
        return None, '报名记录不存在'
    if enrollment.token_expires_at and enrollment.token_expires_at < get_business_now():
        return None, '链接已过期'
    if enrollment.status != 'pending_info':
        return None, '该报名信息已提交，无需重复填写'

    student_name = (data.get('name') or '').strip()
    phone = (data.get('phone') or '').strip()
    if not student_name:
        return None, '姓名为必填项'
    if not phone:
        return None, '手机号为必填项'
    try:
        delivery_preference = _normalize_delivery_preference(
            data.get('delivery_preference'),
            required=True,
        )
    except ValueError as exc:
        return None, str(exc)
    availability_intake, intake_errors = preview_availability_intake(data)
    if intake_errors:
        return None, '；'.join(intake_errors)
    normalized_data = dict(data)
    normalized_data['available_slots'] = availability_intake.get('weekly_slots') or []
    normalized_data['excluded_dates'] = availability_intake.get('excluded_dates') or []

    student_user, account_info = _resolve_or_create_student_account(enrollment, student_name, phone)
    profile = student_user.student_profile if student_user and student_user.student_profile else None
    if not profile:
        profile = StudentProfile(user_id=student_user.id if student_user else None, name=student_name, phone=phone)
        db.session.add(profile)

    _apply_student_profile_fields(profile, normalized_data, preserve_missing=True)
    db.session.flush()
    _sync_student_user(student_user, student_name, phone)

    if student_name != enrollment.student_name:
        enrollment.student_name = student_name
    enrollment.student_profile_id = profile.id
    enrollment.delivery_preference = delivery_preference
    enrollment.status = 'pending_schedule'
    _set_enrollment_ai_scheduling_state(enrollment, availability_intake=availability_intake)
    refresh_enrollment_scheduling_ai_state(enrollment)
    db.session.commit()

    return {
        'account': account_info,
        'profile': profile.to_dict(),
        'enrollment': build_enrollment_payload(enrollment),
    }, None


def update_enrollment_intake(enrollment, data):
    """学生本人或教务修改已提交的 intake 信息，并强制回到待排课。"""
    from modules.auth.models import StudentProfile
    from modules.auth.workflow_services import refresh_enrollment_replan_workflows

    if not enrollment:
        return None, '报名记录不存在'
    if enrollment.status not in {'pending_schedule', 'pending_student_confirm'}:
        return None, '当前状态不允许修改学生信息'

    student_name = (data.get('name') or enrollment.student_name or '').strip()
    phone = (data.get('phone') or '').strip()
    if not student_name:
        return None, '姓名为必填项'
    if not phone:
        return None, '手机号为必填项'
    try:
        delivery_preference = _normalize_delivery_preference(
            data.get('delivery_preference'),
            required=True,
        )
    except ValueError as exc:
        return None, str(exc)
    availability_intake, intake_errors = preview_availability_intake(data)
    if intake_errors:
        return None, '；'.join(intake_errors)
    normalized_data = dict(data)
    normalized_data['available_slots'] = availability_intake.get('weekly_slots') or []
    normalized_data['excluded_dates'] = availability_intake.get('excluded_dates') or []

    student_user, account_info = _resolve_or_create_student_account(enrollment, student_name, phone)
    profile = enrollment.student_profile or (student_user.student_profile if student_user else None)
    if not profile:
        profile = StudentProfile(user_id=student_user.id if student_user else None, name=student_name, phone=phone)
        db.session.add(profile)

    _apply_student_profile_fields(profile, normalized_data, preserve_missing=True)
    db.session.flush()
    _sync_student_user(student_user, student_name, phone)

    enrollment.student_name = student_name
    enrollment.student_profile_id = profile.id
    enrollment.delivery_preference = delivery_preference
    previous_confirmed_slot = None
    if enrollment.confirmed_slot:
        try:
            previous_confirmed_slot = normalize_plan(json.loads(enrollment.confirmed_slot), enrollment)
        except (json.JSONDecodeError, TypeError):
            previous_confirmed_slot = None
    enrollment.proposed_slots = None
    enrollment.confirmed_slot = None
    enrollment.status = 'pending_schedule'
    _set_enrollment_ai_scheduling_state(enrollment, availability_intake=availability_intake)
    refresh_enrollment_scheduling_ai_state(enrollment)
    refresh_enrollment_replan_workflows(
        enrollment,
        reset_to_teacher_proposal=True,
        reason='学生信息已更新，请基于最新可上课时间重新提案。',
        previous_confirmed_slot=previous_confirmed_slot,
    )
    db.session.commit()

    return {
        'account': account_info,
        'profile': profile.to_dict(),
        'enrollment': build_enrollment_payload(enrollment),
    }, None


def save_student_profile_record(data, profile=None):
    """创建或更新学生档案。"""
    from modules.auth.models import Enrollment, StudentProfile, User

    student_name = (data.get('name') or (profile.name if profile else '') or '').strip()
    if not student_name:
        return None, None, '缺少必填字段: name'

    phone = (data.get('phone') or '').strip() or None
    linked_user = None
    account_info = None

    user_id = data.get('user_id')
    username = (data.get('username') or '').strip()
    create_user_if_missing = _coerce_bool(data.get('create_user_if_missing'))

    if user_id:
        linked_user = db.session.get(User, user_id)
        if not linked_user:
            return None, None, '用户不存在'
    elif username:
        linked_user = User.query.filter_by(username=username).first()
        if not linked_user and create_user_if_missing:
            linked_user = User(
                username=username,
                display_name=student_name,
                role='student',
                phone=phone,
            )
            linked_user.set_password(data.get('password') or 'scf123')
            db.session.add(linked_user)
            db.session.flush()
            account_info = {
                'username': linked_user.username,
                'password': data.get('password') or 'scf123',
                'user_id': linked_user.id,
            }

    if profile is None:
        existing_profile = linked_user.student_profile if linked_user and linked_user.student_profile else None
        profile = existing_profile or StudentProfile(user_id=linked_user.id if linked_user else None, name=student_name)
        if existing_profile is None:
            db.session.add(profile)
    elif linked_user and profile.user_id not in (None, linked_user.id):
        return None, None, '该档案已绑定其他用户'

    if linked_user and profile.user_id is None:
        profile.user_id = linked_user.id

    _apply_student_profile_fields(profile, data, preserve_missing=profile.id is not None)
    db.session.flush()

    enrollment_id = data.get('enrollment_id')
    if enrollment_id:
        enrollment = db.session.get(Enrollment, enrollment_id)
        if not enrollment:
            return None, None, '报名记录不存在'
        enrollment.student_profile_id = profile.id
        if student_name:
            enrollment.student_name = student_name

    db.session.commit()
    return profile, account_info, None


def reject_enrollment_schedule(enrollment_id, message_text, actor_user_id=None):
    """外部接口或脚本退回排课方案。"""
    from modules.auth.models import ChatMessage, Enrollment, User
    from modules.auth.workflow_services import ensure_enrollment_replan_workflow

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return False, '报名记录不存在'
    if enrollment.status != 'pending_student_confirm':
        return False, '当前没有待退回的排课方案'

    sender_id = actor_user_id
    if sender_id is None and enrollment.student_profile:
        sender_id = enrollment.student_profile.user_id
    if sender_id is None:
        return False, '无法确定反馈发送人'

    msg = ChatMessage(
        sender_id=sender_id,
        receiver_id=enrollment.teacher_id,
        enrollment_id=enrollment.id,
        content=f'{FEEDBACK_PREFIX}[{enrollment.course_name}] {message_text or "学生对排课方案有疑问，请查看。"}',
        is_read=False,
    )
    db.session.add(msg)
    ensure_enrollment_replan_workflow(
        enrollment,
        rejection_text=message_text,
        actor_user=db.session.get(User, sender_id),
    )
    enrollment.confirmed_slot = None
    enrollment.status = 'pending_schedule'
    db.session.commit()
    return True, '已发送消息给老师'


def _time_to_minutes(t):
    """'14:00' -> 840"""
    h, m = t.split(':')
    return int(h) * 60 + int(m)


def _minutes_to_time(minutes):
    """840 -> '14:00'"""
    return f'{minutes // 60:02d}:{minutes % 60:02d}'


def _compute_overlap(a_start, a_end, b_start, b_end, min_duration_minutes):
    """计算两个时间段的重叠区间，重叠时长必须 >= min_duration_minutes。"""
    start = max(_time_to_minutes(a_start), _time_to_minutes(b_start))
    end = min(_time_to_minutes(a_end), _time_to_minutes(b_end))
    if end - start >= min_duration_minutes:
        return (_minutes_to_time(start), _minutes_to_time(end))
    return None


def _slot_signature(slot):
    return (slot['day_of_week'], slot['time_start'], slot['time_end'])


def _slot_sort_key(slot):
    return (
        slot['day_of_week'],
        _time_to_minutes(slot['time_start']),
        _time_to_minutes(slot['time_end']),
    )


def _first_occurrence_on_or_after(day_of_week, start_date):
    days_ahead = day_of_week - start_date.weekday()
    if days_ahead < 0:
        days_ahead += 7
    return start_date + timedelta(days=days_ahead)


def _get_total_sessions(enrollment):
    if enrollment.total_hours and enrollment.hours_per_session:
        return max(1, int(enrollment.total_hours / enrollment.hours_per_session))
    return 16


def _load_student_excluded_dates(student_profile):
    excluded_set = set()
    if student_profile and student_profile.excluded_dates:
        try:
            excluded_set = set(json.loads(student_profile.excluded_dates) or [])
        except (json.JSONDecodeError, TypeError):
            excluded_set = set()
    return excluded_set


def _load_json_list(value):
    if value in (None, ''):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            data = json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return []
        return data if isinstance(data, list) else []
    return []


def _day_name(day_of_week):
    labels = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    try:
        index = int(day_of_week)
    except (TypeError, ValueError):
        return ''
    if 0 <= index < len(labels):
        return labels[index]
    return ''


def _is_valid_time_text(value):
    return bool(re.match(r'^(?:[01]\d|2[0-3]):[0-5]\d$', str(value or '').strip()))


def _validate_available_slot_entries(value):
    if value in (None, ''):
        return [], []

    if isinstance(value, str):
        try:
            raw_items = json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return [], ['可补课时段格式错误，请传数组']
    else:
        raw_items = value
    if not isinstance(raw_items, list):
        return [], ['可补课时段格式错误，请传数组']

    normalized = []
    errors = []
    for index, slot in enumerate(raw_items, start=1):
        if not isinstance(slot, dict):
            errors.append(f'第 {index} 个可补课时段格式不正确')
            continue
        day_value = slot.get('day')
        if day_value is None:
            day_value = slot.get('day_of_week')
        try:
            day = int(day_value)
        except (TypeError, ValueError):
            errors.append(f'第 {index} 个可补课时段缺少有效星期')
            continue
        if day < 0 or day > 6:
            errors.append(f'第 {index} 个可补课时段星期超出范围')
            continue
        start = (slot.get('start') or slot.get('time_start') or '').strip()
        end = (slot.get('end') or slot.get('time_end') or '').strip()
        if not _is_valid_time_text(start) or not _is_valid_time_text(end):
            errors.append(f'第 {index} 个可补课时段时间格式错误，请使用 HH:MM')
            continue
        if end <= start:
            errors.append(f'第 {index} 个可补课时段结束时间必须晚于开始时间')
            continue
        normalized.append({
            'day': day,
            'start': start,
            'end': end,
        })

    deduped = list({
        (item['day'], item['start'], item['end']): item
        for item in normalized
    }.values())
    deduped.sort(key=lambda item: (item['day'], item['start'], item['end']))
    return deduped, errors


def _normalize_available_slot_entries(value):
    normalized, _ = _validate_available_slot_entries(value)
    return normalized


def _validate_excluded_dates_entries(value):
    if value in (None, ''):
        return [], []

    if isinstance(value, str):
        try:
            raw_items = json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return [], ['禁排日期格式错误，请传数组']
    else:
        raw_items = value
    if not isinstance(raw_items, list):
        return [], ['禁排日期格式错误，请传数组']

    normalized = []
    errors = []
    for index, item in enumerate(raw_items, start=1):
        date_text = str(item or '').strip()
        if not date_text:
            errors.append(f'第 {index} 个禁排日期为空')
            continue
        try:
            date.fromisoformat(date_text)
        except ValueError:
            errors.append(f'第 {index} 个禁排日期格式错误，请使用 YYYY-MM-DD')
            continue
        normalized.append(date_text)
    return sorted(dict.fromkeys(normalized)), errors


def _normalize_excluded_dates_entries(value):
    normalized, _ = _validate_excluded_dates_entries(value)
    return normalized


def _summarize_available_slots(value):
    slots = _normalize_available_slot_entries(value)
    if not slots:
        return None
    return '；'.join(
        f'{_day_name(slot["day"])} {slot["start"]}-{slot["end"]}'
        for slot in slots
        if _day_name(slot['day'])
    ) or None


def _summarize_excluded_dates(value):
    dates = _normalize_excluded_dates_entries(value)
    if not dates:
        return None
    if len(dates) <= 3:
        return '、'.join(dates)
    return f'{dates[0]}、{dates[1]} 等 {len(dates)} 天'


def _summarize_schedule_summary(schedule_data):
    if not schedule_data:
        return None

    if isinstance(schedule_data, dict):
        day_of_week = schedule_data.get('day_of_week')
        date_text = schedule_data.get('date')
        time_start = schedule_data.get('time_start')
        time_end = schedule_data.get('time_end')
    else:
        day_of_week = getattr(schedule_data, 'day_of_week', None)
        schedule_date = getattr(schedule_data, 'date', None)
        date_text = schedule_date.isoformat() if schedule_date else None
        time_start = getattr(schedule_data, 'time_start', None)
        time_end = getattr(schedule_data, 'time_end', None)

    parts = []
    day_label = _day_name(day_of_week)
    if day_label:
        parts.append(day_label)
    if date_text:
        parts.append(date_text)
    if time_start and time_end:
        parts.append(f'{time_start}-{time_end}')
    return ' '.join(parts) or None


def _summarize_plan(plan):
    if not isinstance(plan, dict):
        return None

    weekly_slots = plan.get('weekly_slots') or []
    session_dates = plan.get('session_dates') or []
    parts = []

    weekly_summary = _summarize_available_slots([
        {
            'day': slot.get('day_of_week'),
            'start': slot.get('time_start'),
            'end': slot.get('time_end'),
        }
        for slot in weekly_slots
    ])
    if weekly_summary:
        parts.append(f'每周 {weekly_summary}')

    if session_dates:
        parts.append(f'共 {len(session_dates)} 节')
        first_summary = _summarize_schedule_summary(session_dates[0])
        last_summary = _summarize_schedule_summary(session_dates[-1])
        if first_summary and last_summary:
            parts.append(first_summary if first_summary == last_summary else f'{first_summary} 至 {last_summary}')

    return ' · '.join(parts) or None


def _summarize_makeup_preferences(available_slots, excluded_dates, note=None):
    parts = []
    slot_summary = _summarize_available_slots(available_slots)
    if slot_summary:
        parts.append(f'本次可补课时间：{slot_summary}')
    excluded_summary = _summarize_excluded_dates(excluded_dates)
    if excluded_summary:
        parts.append(f'禁排日期：{excluded_summary}')
    note_text = (note or '').strip()
    if note_text:
        parts.append(f'补课备注：{note_text}')
    return ' · '.join(parts) or None


def _session_preview_lines(session_dates, *, limit=3):
    normalized_dates, errors = _normalize_manual_session_dates(session_dates or [])
    source = normalized_dates if not errors else []
    lines = [
        f'{item["date"]} {item["time_start"]}-{item["time_end"]}'
        for item in source[:limit]
    ]
    hidden_count = max(len(source) - limit, 0)
    if hidden_count:
        lines.append(f'其余 {hidden_count} 节待查看详情')
    return lines


def _profile_constraint_meta(student_profile):
    available_slots = _normalize_available_slot_entries(
        student_profile.available_slots if student_profile else None
    )
    excluded_dates = _normalize_excluded_dates_entries(
        student_profile.excluded_dates if student_profile else None
    )
    notes = (getattr(student_profile, 'notes', None) or '').strip() or None
    return {
        'available_times_summary': _summarize_available_slots(available_slots),
        'excluded_dates_summary': _summarize_excluded_dates(excluded_dates),
        'available_slots': available_slots,
        'excluded_dates': excluded_dates,
        'notes': notes,
    }


def _build_scheduling_complexity_hint(enrollment):
    profile = getattr(enrollment, 'student_profile', None)
    meta = _profile_constraint_meta(profile)
    flags = []
    if not meta['available_slots']:
        flags.append('学生还没有长期可上课时间，建议先补充后再排课')
    elif len(meta['available_slots']) <= 1:
        flags.append('学生长期可上课时段较少，排课弹性有限')
    if len(meta['excluded_dates']) >= 3:
        flags.append('学生禁排日期较多，建议先核对本周安排')
    if (enrollment.total_hours or 0) >= 10:
        flags.append('总课时较多，建议优先选择稳定可复用时段')
    if meta['notes']:
        flags.append('报名备注里有额外限制，排课前请先阅读')
    if not flags:
        return '当前排课约束较少，可直接进入排课'
    return '；'.join(flags[:3])


def _availability_input_text(data):
    return (
        data.get('availability_input_text')
        or data.get('availability_text')
        or data.get('time_preference_text')
        or ''
    )


def _availability_evidence_items(data):
    evidence_items = data.get('availability_evidence_items') or []
    evidence_text = (data.get('availability_evidence_text') or '').strip()
    if evidence_text and not any(
        isinstance(item, dict)
        and str(item.get('type') or '').strip().lower() == 'text_capture'
        and str(item.get('content') or item.get('text') or '').strip() == evidence_text
        for item in evidence_items
    ):
        evidence_items = [*evidence_items, {'type': 'text_capture', 'content': evidence_text}]
    return resolve_availability_evidence_items(evidence_items)


def _availability_confirmed_parse_result(data):
    confirmed = data.get('confirmed_parse_result')
    return confirmed if isinstance(confirmed, dict) else {}


def preview_availability_intake(data):
    confirmed_parse = _availability_confirmed_parse_result(data)
    manual_adjustments = data.get('manual_adjustments') or {}
    raw_evidence_items = data.get('availability_evidence_items') or []
    manual_slot_source = (
        data.get('available_times')
        if 'available_times' in data
        else data.get('available_slots')
    )
    if manual_slot_source in (None, '', []) and isinstance(manual_adjustments, dict):
        manual_slot_source = manual_adjustments.get('weekly_slots')
    manual_slots, slot_errors = _validate_available_slot_entries(
        manual_slot_source
    )
    manual_excluded_source = data.get('excluded_dates')
    if manual_excluded_source in (None, '', []) and isinstance(manual_adjustments, dict):
        manual_excluded_source = manual_adjustments.get('excluded_dates')
    manual_excluded_dates, excluded_errors = _validate_excluded_dates_entries(manual_excluded_source)
    errors = slot_errors + excluded_errors
    availability_input_text = _availability_input_text(data)
    resolved_evidence_items = _availability_evidence_items(data)
    parsed = parse_availability_intake(
        input_text=availability_input_text,
        evidence_items=resolved_evidence_items,
        manual_slots=manual_slots,
        manual_excluded_dates=manual_excluded_dates,
        reference_date=get_business_today(),
    )
    confirmed_slots = _normalize_available_slot_entries(confirmed_parse.get('weekly_slots'))
    confirmed_excluded_dates = _normalize_excluded_dates_entries(confirmed_parse.get('excluded_dates'))
    resolved_slots = manual_slots or confirmed_slots or parsed.get('weekly_slots') or []
    resolved_excluded_dates = manual_excluded_dates or confirmed_excluded_dates or parsed.get('excluded_dates') or []
    if not resolved_slots:
        has_image_evidence = any(
            isinstance(item, dict)
            and str(item.get('type') or '').strip().lower() in {'image', 'image_url', 'image_data_url', 'image_base64'}
            for item in raw_evidence_items
        )
        if has_image_evidence and not availability_input_text.strip() and not resolved_evidence_items:
            errors.append('图片还没有识别出可用时间，请补一段文字说明，或检查豆包视觉配置后重试')
        else:
            errors.append('请先输入可上课时间，或至少确认一个结构化时段')

    intake = {
        **confirmed_parse,
        **parsed,
        'weekly_slots': resolved_slots,
        'excluded_dates': resolved_excluded_dates,
        'summary': build_availability_intake_summary({
            'weekly_slots': resolved_slots,
            'excluded_dates': resolved_excluded_dates,
            'temporary_constraints': parsed.get('temporary_constraints') or [],
        }),
    }
    return intake, errors


def _set_enrollment_ai_scheduling_state(
    enrollment,
    *,
    availability_intake=None,
    candidate_slot_pool=None,
    recommended_bundle=None,
    risk_assessment=None,
):
    if availability_intake is not None:
        enrollment.availability_intake = _serialize_json_field(availability_intake)
    if candidate_slot_pool is not None:
        enrollment.candidate_slot_pool = _serialize_json_field(candidate_slot_pool)
    if recommended_bundle is not None:
        enrollment.recommended_bundle = _serialize_json_field(recommended_bundle)
    if risk_assessment is not None:
        enrollment.risk_assessment = _serialize_json_field(risk_assessment)


def _enrollment_availability_intake_meta(enrollment):
    intake = _enrollment_json_field(getattr(enrollment, 'availability_intake', None)) or {}
    if intake:
        return intake
    profile = getattr(enrollment, 'student_profile', None)
    meta = _profile_constraint_meta(profile)
    return {
        'weekly_slots': meta.get('available_slots') or [],
        'excluded_dates': meta.get('excluded_dates') or [],
        'temporary_constraints': [],
        'confidence': 1.0 if meta.get('available_slots') else 0.0,
        'needs_review': False,
        'summary': build_availability_intake_summary({
            'weekly_slots': meta.get('available_slots') or [],
            'excluded_dates': meta.get('excluded_dates') or [],
            'temporary_constraints': [],
        }),
    }


def _student_available_slots_for_enrollment(enrollment):
    intake_meta = _enrollment_availability_intake_meta(enrollment)
    intake_slots = intake_meta.get('weekly_slots') or []
    if intake_slots:
        return [
            {
                'day_of_week': item['day'],
                'time_start': item['start'],
                'time_end': item['end'],
            }
            for item in _normalize_available_slot_entries(intake_slots)
        ]
    return _load_student_available_ranges(enrollment.student_profile)


def _target_finish_date_summary(target_finish_date):
    return target_finish_date.isoformat() if target_finish_date else None


def _candidate_pool_label(candidate):
    return f'{_day_name(candidate["day_of_week"])} {candidate["time_start"]}-{candidate["time_end"]}'


def _candidate_pool_for_enrollment(enrollment):
    from modules.oa.models import CourseSchedule

    teacher_context = _teacher_work_context(enrollment.teacher or enrollment.teacher_id)
    teacher_slots = _load_teacher_available_ranges(enrollment.teacher_id)
    student_slots = _student_available_slots_for_enrollment(enrollment)
    if not student_slots:
        return [], teacher_slots, {
            **teacher_context,
            'teacher_availability_ready': bool(teacher_slots),
        }

    if not teacher_slots:
        return [], teacher_slots, {
            **teacher_context,
            'teacher_availability_ready': False,
        }

    min_minutes = max(int((enrollment.hours_per_session or 2.0) * 60), 30)
    query = CourseSchedule.query.filter(
        CourseSchedule.teacher_id == enrollment.teacher_id,
        CourseSchedule.is_cancelled == False,
    )
    if enrollment.id:
        query = query.filter(
            or_(CourseSchedule.enrollment_id.is_(None), CourseSchedule.enrollment_id != enrollment.id)
        )

    existing_by_day = {}
    for schedule in query.all():
        existing_by_day.setdefault(schedule.day_of_week, []).append(schedule)

    candidates = []
    for teacher_slot in teacher_slots:
        teacher_day = int(teacher_slot.get('day_of_week', teacher_slot.get('day', -1)))
        teacher_start = teacher_slot.get('time_start', teacher_slot.get('start', ''))
        teacher_end = teacher_slot.get('time_end', teacher_slot.get('end', ''))
        for student_slot in student_slots:
            if int(student_slot.get('day', student_slot.get('day_of_week', -1))) != teacher_day:
                continue

            overlap = _compute_overlap(
                teacher_start,
                teacher_end,
                student_slot.get('start', student_slot.get('time_start', '')),
                student_slot.get('end', student_slot.get('time_end', '')),
                min_minutes,
            )
            if not overlap:
                continue

            overlap_start = _time_to_minutes(overlap[0])
            overlap_end = _time_to_minutes(overlap[1])
            cursor = overlap_start
            while cursor + min_minutes <= overlap_end:
                block_start = _minutes_to_time(cursor)
                block_end = _minutes_to_time(cursor + min_minutes)

                conflicts = []
                for existing in existing_by_day.get(teacher_day, []):
                    if _compute_overlap(
                        block_start,
                        block_end,
                        existing.time_start,
                        existing.time_end,
                        1,
                    ):
                        conflicts.append({
                            'course_name': existing.course_name,
                            'time': f'{existing.time_start}-{existing.time_end}',
                            'students': existing.students,
                        })

                score = 0
                if not conflicts:
                    score += 4
                if teacher_slot.get('is_preferred'):
                    score += 1
                if block_start >= student_slot.get('start', student_slot.get('time_start', '')) and block_end <= student_slot.get('end', student_slot.get('time_end', '')):
                    score += 2

                candidates.append({
                    'day_of_week': teacher_slot['day_of_week'],
                    'time_start': block_start,
                    'time_end': block_end,
                    'score': score,
                    'is_preferred': bool(teacher_slot.get('is_preferred')),
                    'conflicts': conflicts,
                    'label': f'{_day_name(teacher_slot["day_of_week"])} {block_start}-{block_end}',
                })
                cursor += 60

    deduped = {}
    for candidate in candidates:
        key = _slot_signature(candidate)
        if key not in deduped or candidate.get('score', 0) > deduped[key].get('score', 0):
            deduped[key] = candidate

    candidate_pool = [
        {
            **candidate,
            'label': _candidate_pool_label(candidate),
            'within_student_constraints': True,
        }
        for candidate in sorted(deduped.values(), key=_candidate_sort_key)
        if not candidate.get('conflicts')
    ]
    return candidate_pool, teacher_slots, {
        **teacher_context,
        'teacher_availability_ready': bool(teacher_slots),
    }


def _assess_enrollment_scheduling_risk(enrollment, candidate_pool, recommended_bundle):
    required_weekly = max(int(enrollment.sessions_per_week or 1), 1)
    intake_meta = _enrollment_availability_intake_meta(enrollment)
    teacher_context = _teacher_work_context(enrollment.teacher or enrollment.teacher_id)
    hard_errors = []
    warnings = []
    clarification_reasons = []
    teacher_confirmation_required = False

    if not _student_available_slots_for_enrollment(enrollment):
        clarification_reasons.append('学生当前还没有足够明确的可上课时间，需先补充时间偏好')

    teacher_ranges = _load_teacher_available_ranges(enrollment.teacher_id)
    if not teacher_ranges:
        if teacher_context['teacher_work_mode'] == TEACHER_WORK_MODE_FULL_TIME:
            hard_errors.append('全职老师默认工作模板暂不可用，需教务先处理老师排课模板')
        else:
            hard_errors.append('兼职老师还没有维护长期 availability，暂不能进入老师协同排课')

    if len(candidate_pool) < required_weekly:
        if teacher_context['teacher_work_mode'] == TEACHER_WORK_MODE_FULL_TIME and _student_available_slots_for_enrollment(enrollment):
            teacher_confirmation_required = True
            warnings.append(
                f'学生时间无法完全落在全职老师默认工作模板内，如需满足每周 {required_weekly} 节，必须由老师确认模板外时段'
            )
        else:
            hard_errors.append(f'学生当前可排时段不足以支撑每周 {required_weekly} 节，需教务介入补齐时间')

    if recommended_bundle:
        target_finish_date = getattr(enrollment, 'target_finish_date', None)
        if target_finish_date and recommended_bundle.get('date_end'):
            try:
                bundle_end = date.fromisoformat(recommended_bundle['date_end'])
            except ValueError:
                bundle_end = None
            if bundle_end and bundle_end > target_finish_date:
                hard_errors.append(
                    f'按当前推荐方案预计到 {bundle_end.isoformat()} 才能完成，晚于目标完成日 {target_finish_date.isoformat()}'
                )

    confidence = float(intake_meta.get('confidence') or 0)
    if confidence and confidence < 0.75:
        clarification_reasons.append('学生时间解析置信度较低，建议让学生再确认一次结果')
    elif intake_meta.get('needs_review'):
        clarification_reasons.append('学生时间输入仍有待确认的细节，建议先让学生补充说明')
    if candidate_pool and len(candidate_pool) <= required_weekly:
        warnings.append('当前候选时间刚好满足每周配额，方案弹性较低')
    elif candidate_pool and len(candidate_pool) <= required_weekly + 1:
        warnings.append('当前候选时间较少，如后续再冲突可能需要教务介入')
    if intake_meta.get('temporary_constraints'):
        warnings.append('学生输入里包含临时限制，建议老师提交前先核对上下文')

    recommended_action = 'direct_to_student'
    if clarification_reasons:
        recommended_action = 'needs_student_clarification'
    elif hard_errors:
        recommended_action = 'needs_admin_intervention'
    elif teacher_confirmation_required:
        recommended_action = 'needs_teacher_confirmation'
    elif warnings:
        recommended_action = 'needs_admin_review'

    summary = None
    if clarification_reasons:
        summary = clarification_reasons[0]
    elif hard_errors:
        summary = hard_errors[0]
    elif teacher_confirmation_required:
        summary = warnings[0] if warnings else '当前方案需要老师确认模板外时段'
    elif warnings:
        summary = warnings[0]
    elif recommended_bundle:
        summary = f'已找到符合每周 {required_weekly} 节的推荐方案，可直接发给学生确认'

    return {
        'hard_errors': list(dict.fromkeys(hard_errors)),
        'warnings': list(dict.fromkeys(warnings)),
        'clarification_reasons': list(dict.fromkeys(clarification_reasons)),
        'confidence': confidence,
        'needs_review': bool(intake_meta.get('needs_review')),
        'teacher_work_mode': teacher_context['teacher_work_mode'],
        'teacher_work_mode_label': teacher_context['teacher_work_mode_label'],
        'availability_source': teacher_context['availability_source'],
        'teacher_confirmation_required': teacher_confirmation_required,
        'student_clarification_required': bool(clarification_reasons),
        'coverage_gap': {
            'weekly_required': required_weekly,
            'weekly_available': len(candidate_pool),
            'missing': max(required_weekly - len(candidate_pool), 0),
        },
        'recommended_action': recommended_action,
        'severity': 'hard' if hard_errors else ('warning' if (clarification_reasons or warnings) else 'ok'),
        'summary': summary,
    }


def refresh_enrollment_scheduling_ai_state(enrollment):
    if not enrollment:
        return {'candidate_slot_pool': [], 'recommended_bundle': None, 'risk_assessment': None, 'proposed_plans': []}

    candidate_pool, _, _ = _candidate_pool_for_enrollment(enrollment)
    required_weekly = max(int(enrollment.sessions_per_week or 1), 1)
    total_sessions = _get_total_sessions(enrollment)
    excluded_set = _load_student_excluded_dates(enrollment.student_profile)

    top_candidates = candidate_pool[: min(len(candidate_pool), max(required_weekly + 6, 8))]
    valid_plans = []
    for combo in combinations(top_candidates, required_weekly):
        selected_blocks = sorted(list(combo), key=_slot_sort_key)
        plan = _build_plan(selected_blocks, total_sessions, excluded_set)
        normalized_plan = normalize_plan(plan, enrollment)
        plan_errors, _ = _collect_manual_plan_issues(
            enrollment,
            normalized_plan.get('session_dates', []),
            weekly_slots=normalized_plan.get('weekly_slots') or selected_blocks,
        )
        target_finish_date = getattr(enrollment, 'target_finish_date', None)
        if target_finish_date and normalized_plan.get('date_end'):
            try:
                plan_end = date.fromisoformat(normalized_plan['date_end'])
            except ValueError:
                plan_end = None
            if plan_end and plan_end > target_finish_date:
                plan_errors.append(
                    f'方案完成时间晚于目标完成日 {target_finish_date.isoformat()}'
                )
        if plan_errors:
            continue
        valid_plans.append(normalized_plan)
        if len(valid_plans) >= 3:
            break

    recommended_bundle = valid_plans[0] if valid_plans else None
    risk_assessment = _assess_enrollment_scheduling_risk(enrollment, candidate_pool, recommended_bundle)
    _set_enrollment_ai_scheduling_state(
        enrollment,
        candidate_slot_pool=candidate_pool,
        recommended_bundle=recommended_bundle,
        risk_assessment=risk_assessment,
    )
    return {
        'candidate_slot_pool': candidate_pool,
        'recommended_bundle': recommended_bundle,
        'risk_assessment': risk_assessment,
        'proposed_plans': valid_plans,
    }


def _build_session_dates(weekly_slots, total_sessions, excluded_set):
    """按周轮转生成具体课次，遇到不可上课日期则跳过并顺延到下周。"""
    if not weekly_slots:
        return [], []

    sorted_slots = sorted(weekly_slots, key=_slot_sort_key)
    today = get_business_today()
    first_dates = [
        _first_occurrence_on_or_after(slot['day_of_week'], today)
        for slot in sorted_slots
    ]

    session_dates = []
    skipped_dates = []
    week_idx = 0
    max_weeks = total_sessions + len(excluded_set) + 52

    while len(session_dates) < total_sessions and week_idx < max_weeks:
        for slot, first_date in zip(sorted_slots, first_dates):
            current_date = first_date + timedelta(weeks=week_idx)
            payload = {
                'date': current_date.isoformat(),
                'day_of_week': slot['day_of_week'],
                'time_start': slot['time_start'],
                'time_end': slot['time_end'],
            }
            if payload['date'] in excluded_set:
                skipped_dates.append(payload)
                continue
            session_dates.append(payload)
            if len(session_dates) >= total_sessions:
                break
        week_idx += 1

    return session_dates, skipped_dates


def _build_plan(weekly_slots, total_sessions, excluded_set):
    """把每周 recurring slots 组装成一个可确认的排课 plan。"""
    deduped = {}
    for slot in weekly_slots:
        key = _slot_signature(slot)
        if key not in deduped or slot.get('score', 0) > deduped[key].get('score', 0):
            deduped[key] = dict(slot)

    ordered_slots = sorted(deduped.values(), key=_slot_sort_key)
    session_dates, skipped_dates = _build_session_dates(ordered_slots, total_sessions, excluded_set)
    date_values = [item['date'] for item in session_dates]
    estimated_weeks = 0
    if date_values:
        first_date = date.fromisoformat(date_values[0])
        last_date = date.fromisoformat(date_values[-1])
        estimated_weeks = ((last_date - first_date).days // 7) + 1

    total_conflicts = sum(len(slot.get('conflicts', [])) for slot in ordered_slots)
    plan_score = sum(slot.get('score', 0) for slot in ordered_slots)
    distinct_days = len({slot['day_of_week'] for slot in ordered_slots})

    return {
        'weekly_slots': ordered_slots,
        'total_sessions': total_sessions,
        'sessions_per_week': len(ordered_slots),
        'estimated_weeks': estimated_weeks,
        'session_dates': session_dates,
        'skipped_dates': skipped_dates,
        'plan_score': plan_score,
        'distinct_days': distinct_days,
        'total_conflicts': total_conflicts,
        'date_start': date_values[0] if date_values else None,
        'date_end': date_values[-1] if date_values else None,
    }


def _apply_session_dates_to_plan(plan, session_dates, skipped_dates):
    date_values = [item['date'] for item in session_dates]
    estimated_weeks = 0
    if date_values:
        first_date = date.fromisoformat(date_values[0])
        last_date = date.fromisoformat(date_values[-1])
        estimated_weeks = ((last_date - first_date).days // 7) + 1

    plan['session_dates'] = session_dates
    plan['skipped_dates'] = skipped_dates
    plan['date_start'] = date_values[0] if date_values else None
    plan['date_end'] = date_values[-1] if date_values else None
    plan['estimated_weeks'] = estimated_weeks
    plan['total_sessions'] = len(session_dates) or plan.get('total_sessions', 0)
    plan['sessions_per_week'] = len(plan.get('weekly_slots', []))
    plan['distinct_days'] = len({slot['day_of_week'] for slot in plan.get('weekly_slots', [])})
    plan['total_conflicts'] = sum(len(slot.get('conflicts', [])) for slot in plan.get('weekly_slots', []))
    plan['plan_score'] = sum(slot.get('score', 0) for slot in plan.get('weekly_slots', []))
    return plan


def _merge_plan_metadata(raw_plan, plan):
    if not isinstance(raw_plan, dict):
        return plan

    preserved_keys = {
        'is_manual',
        'manual_note',
        'manual_warnings',
        'note',
        'warnings',
        'submitted_by',
        'submitted_by_name',
        'submitted_at',
        'quota_required',
        'quota_selected',
    }
    for key in preserved_keys:
        if key in raw_plan:
            plan[key] = raw_plan[key]
    return plan


def normalize_plan(raw_plan, enrollment=None):
    """兼容旧单时段结构，统一为新的 multi-slot plan 结构。"""
    if not raw_plan:
        return None
    if not isinstance(raw_plan, dict):
        return raw_plan

    excluded_set = _load_student_excluded_dates(enrollment.student_profile) if enrollment else set()
    total_sessions = _get_total_sessions(enrollment) if enrollment else raw_plan.get('total_sessions', 1)

    if raw_plan.get('weekly_slots'):
        plan = _build_plan(raw_plan.get('weekly_slots', []), total_sessions, excluded_set)
        session_dates = raw_plan.get('session_dates')
        skipped_dates = raw_plan.get('skipped_dates')
        if session_dates is not None:
            plan = _apply_session_dates_to_plan(
                plan,
                session_dates,
                skipped_dates or [],
            )
        return _merge_plan_metadata(raw_plan, plan)

    if all(key in raw_plan for key in ('day_of_week', 'time_start', 'time_end')):
        weekly_slots = [{
            'day_of_week': raw_plan['day_of_week'],
            'time_start': raw_plan['time_start'],
            'time_end': raw_plan['time_end'],
            'score': raw_plan.get('score', 0),
            'is_preferred': raw_plan.get('is_preferred', False),
            'conflicts': raw_plan.get('conflicts', []),
        }]
        plan = _build_plan(weekly_slots, total_sessions, excluded_set)
        if raw_plan.get('dates') is not None:
            session_dates = [{
                'date': item,
                'day_of_week': raw_plan['day_of_week'],
                'time_start': raw_plan['time_start'],
                'time_end': raw_plan['time_end'],
            } for item in raw_plan.get('dates', [])]
            skipped_dates = [{
                'date': item,
                'day_of_week': raw_plan['day_of_week'],
                'time_start': raw_plan['time_start'],
                'time_end': raw_plan['time_end'],
            } for item in raw_plan.get('skipped_dates', [])]
            plan = _apply_session_dates_to_plan(plan, session_dates, skipped_dates)
        return _merge_plan_metadata(raw_plan, plan)

    return raw_plan


def _plan_sort_key(plan):
    return (
        plan['estimated_weeks'] or 999,
        -plan['distinct_days'],
        -plan['plan_score'],
        plan['total_conflicts'],
        tuple(_slot_signature(slot) for slot in plan['weekly_slots']),
    )


def _candidate_sort_key(candidate):
    return (
        -candidate.get('score', 0),
        len(candidate.get('conflicts', [])),
        candidate['day_of_week'],
        _time_to_minutes(candidate['time_start']),
    )


def _extend_single_day_plan(selected_blocks, grouped_candidates, total_sessions):
    """只有单一 weekday 时，允许补一个同日的第二个 recurring block。"""
    if len({block['day_of_week'] for block in selected_blocks}) > 1:
        return selected_blocks

    used = {_slot_signature(block) for block in selected_blocks}
    extras = []
    for blocks in grouped_candidates.values():
        for block in blocks:
            if _slot_signature(block) in used:
                continue
            extras.append(block)

    extras.sort(key=_candidate_sort_key)
    if extras and len(selected_blocks) < total_sessions:
        return sorted(list(selected_blocks) + [extras[0]], key=_slot_sort_key)
    return selected_blocks


def _schedule_start_datetime(schedule):
    start_time = datetime.strptime(schedule.time_start, '%H:%M').time()
    return datetime.combine(schedule.date, start_time)


def _schedule_end_datetime(schedule):
    end_time = datetime.strptime(schedule.time_end, '%H:%M').time()
    return datetime.combine(schedule.date, end_time)


def _schedule_has_started(schedule, reference=None):
    if not schedule or not schedule.date or not schedule.time_start:
        return False
    reference = reference or get_business_now()
    return _schedule_start_datetime(schedule) <= reference


def _parse_session_input(session, index):
    if not isinstance(session, dict):
        return None, f'第 {index} 节课程数据格式不正确'

    session_date = (session.get('date') or '').strip()
    time_start = (session.get('time_start') or '').strip()
    time_end = (session.get('time_end') or '').strip()
    if not session_date or not time_start or not time_end:
        return None, f'第 {index} 节课程缺少日期或时间'

    try:
        session_day = date.fromisoformat(session_date)
    except ValueError:
        return None, f'第 {index} 节课程日期格式无效'

    try:
        start_dt = datetime.strptime(time_start, '%H:%M')
        end_dt = datetime.strptime(time_end, '%H:%M')
    except ValueError:
        return None, f'第 {index} 节课程时间格式无效'

    if end_dt <= start_dt:
        return None, f'第 {index} 节课程结束时间必须晚于开始时间'

    return {
        'date': session_day.isoformat(),
        'day_of_week': session_day.weekday(),
        'time_start': time_start,
        'time_end': time_end,
    }, None


def _normalize_manual_session_dates(session_dates):
    normalized = []
    errors = []
    for index, session in enumerate(session_dates or [], start=1):
        payload, error = _parse_session_input(session, index)
        if error:
            errors.append(error)
            continue
        normalized.append(payload)

    normalized.sort(
        key=lambda item: (
            item['date'],
            _time_to_minutes(item['time_start']),
            _time_to_minutes(item['time_end']),
        )
    )
    return normalized, errors


def _slots_overlap(time_start, time_end, other_start, other_end):
    return _compute_overlap(time_start, time_end, other_start, other_end, 1) is not None


def _session_within_ranges(session, ranges):
    for slot in ranges:
        slot_day = slot.get('day_of_week', slot.get('day'))
        slot_start = slot.get('time_start', slot.get('start'))
        slot_end = slot.get('time_end', slot.get('end'))
        if slot_day != session['day_of_week']:
            continue
        if (
            session['time_start'] >= slot_start
            and session['time_end'] <= slot_end
        ):
            return True
    return False


def _load_student_available_ranges(student_profile):
    if not student_profile or not student_profile.available_slots:
        return []
    try:
        available_slots = json.loads(student_profile.available_slots) or []
    except (json.JSONDecodeError, TypeError):
        return []

    ranges = []
    for slot in available_slots:
        try:
            day = int(slot.get('day'))
        except (TypeError, ValueError):
            continue
        start = (slot.get('start') or '').strip()
        end = (slot.get('end') or '').strip()
        if not start or not end:
            continue
        ranges.append({
            'day_of_week': day,
            'time_start': start,
            'time_end': end,
        })
    return ranges


def _load_teacher_available_ranges(teacher_id):
    from modules.auth.models import TeacherAvailability

    teacher = _resolve_teacher_user_record(teacher_id)
    if not teacher:
        return []
    if resolve_teacher_work_mode(teacher) == TEACHER_WORK_MODE_FULL_TIME:
        return [
            {
                'day_of_week': slot['day'],
                'time_start': slot['start'],
                'time_end': slot['end'],
                'is_preferred': False,
            }
            for slot in resolve_teacher_default_working_template(teacher)
        ]

    slots = TeacherAvailability.query.filter_by(user_id=teacher_id).all()
    return [
        {
            'day_of_week': slot.day_of_week,
            'time_start': slot.time_start,
            'time_end': slot.time_end,
            'is_preferred': bool(slot.is_preferred),
        }
        for slot in slots
    ]


def _actor_can_manage_schedule_feedback(actor, schedule=None, *, teacher_id=None):
    return _schedule_matches_teacher_actor(actor, schedule, teacher_id=teacher_id)


def _collect_teacher_schedule_conflicts(enrollment, session_dates, *, ignore_schedule_ids=None):
    from modules.oa.models import CourseSchedule

    if not enrollment or not session_dates:
        return []

    ignore_ids = set(ignore_schedule_ids or [])
    teacher_name = enrollment.teacher.display_name if enrollment.teacher else ''
    teacher_filters = [CourseSchedule.teacher_id == enrollment.teacher_id]
    if teacher_name:
        teacher_filters.append(CourseSchedule.teacher == teacher_name)

    session_dates_by_date = {}
    for session in session_dates:
        session_dates_by_date.setdefault(session['date'], []).append(session)

    query = CourseSchedule.query.filter(
        CourseSchedule.date.in_([date.fromisoformat(value) for value in session_dates_by_date]),
        or_(*teacher_filters),
        CourseSchedule.is_cancelled == False,
    )
    if ignore_ids:
        query = query.filter(~CourseSchedule.id.in_(ignore_ids))

    conflicts = []
    existing_by_date = {}
    for schedule in query.all():
        existing_by_date.setdefault(schedule.date.isoformat(), []).append(schedule)

    for session in session_dates:
        for existing in existing_by_date.get(session['date'], []):
            if _slots_overlap(
                session['time_start'],
                session['time_end'],
                existing.time_start,
                existing.time_end,
            ):
                conflicts.append(
                    f'{session["date"]} {session["time_start"]}-{session["time_end"]} '
                    f'与老师现有课程冲突：{existing.course_name} {existing.time_start}-{existing.time_end}'
                )

    return conflicts


def _collect_student_schedule_conflicts(enrollment, session_dates, *, ignore_schedule_ids=None):
    from modules.oa.models import CourseSchedule

    if not enrollment or not enrollment.student_profile_id or not session_dates:
        return []

    ignore_ids = set(ignore_schedule_ids or [])
    session_dates_by_date = {}
    for session in session_dates:
        session_dates_by_date.setdefault(session['date'], []).append(session)

    query = CourseSchedule.query.filter(
        student_schedule_profile_clause(enrollment.student_profile_id, schedule_model=CourseSchedule),
        CourseSchedule.date.in_([date.fromisoformat(value) for value in session_dates_by_date]),
        CourseSchedule.is_cancelled == False,
    )
    if ignore_ids:
        query = query.filter(~CourseSchedule.id.in_(ignore_ids))
    if enrollment.id:
        query = query.filter(
            or_(CourseSchedule.enrollment_id.is_(None), CourseSchedule.enrollment_id != enrollment.id)
        )

    conflicts = []
    existing_by_date = {}
    for schedule in query.all():
        existing_by_date.setdefault(schedule.date.isoformat(), []).append(schedule)

    for session in session_dates:
        for existing in existing_by_date.get(session['date'], []):
            if _slots_overlap(
                session['time_start'],
                session['time_end'],
                existing.time_start,
                existing.time_end,
            ):
                conflicts.append(
                    f'{session["date"]} {session["time_start"]}-{session["time_end"]} '
                    f'与同一学生现有课程冲突：{existing.course_name} {existing.time_start}-{existing.time_end}'
                )

    return conflicts


def _build_weekly_slots_from_sessions(session_dates):
    weekly_slots = {}
    for session in session_dates:
        key = (
            session['day_of_week'],
            session['time_start'],
            session['time_end'],
        )
        if key not in weekly_slots:
            weekly_slots[key] = {
                'day_of_week': session['day_of_week'],
                'time_start': session['time_start'],
                'time_end': session['time_end'],
                'score': 0,
                'is_preferred': False,
                'conflicts': [],
            }
    return sorted(weekly_slots.values(), key=_slot_sort_key)


def _build_manual_plan(session_dates):
    weekly_slots = _build_weekly_slots_from_sessions(session_dates)
    plan = _build_plan(weekly_slots, len(session_dates), set())
    plan = _apply_session_dates_to_plan(plan, session_dates, [])
    plan['is_manual'] = True
    return plan


def _normalize_weekly_slots_for_validation(weekly_slots):
    normalized = []
    for slot in weekly_slots or []:
        if not isinstance(slot, dict):
            continue
        try:
            day_of_week = int(slot.get('day_of_week', slot.get('day')))
        except (TypeError, ValueError):
            continue
        time_start = (slot.get('time_start') or slot.get('start') or '').strip()
        time_end = (slot.get('time_end') or slot.get('end') or '').strip()
        if not time_start or not time_end:
            continue
        normalized.append({
            'day_of_week': day_of_week,
            'time_start': time_start,
            'time_end': time_end,
        })
    deduped = {
        _slot_signature(slot): slot
        for slot in normalized
    }
    return sorted(deduped.values(), key=_slot_sort_key)


def _collect_manual_plan_issues(enrollment, session_dates, weekly_slots=None):
    errors = []
    warnings = []

    expected_sessions = _get_total_sessions(enrollment)
    if len(session_dates) != expected_sessions:
        errors.append(f'课次数量必须为 {expected_sessions} 节，当前为 {len(session_dates)} 节')

    if not session_dates:
        errors.append('至少需要保留一节课程')
        return errors, warnings

    plan = _build_manual_plan(session_dates)
    required_weekly = max(int(getattr(enrollment, 'sessions_per_week', 1) or 1), 1)
    selected_weekly_slots = (
        _normalize_weekly_slots_for_validation(weekly_slots)
        if weekly_slots is not None
        else (plan.get('weekly_slots') or [])
    )
    selected_weekly = len(selected_weekly_slots)
    if selected_weekly != required_weekly:
        errors.append(f'每周必须选满 {required_weekly} 节，当前方案只覆盖 {selected_weekly} 节')

    target_finish_date = getattr(enrollment, 'target_finish_date', None)
    if target_finish_date and plan.get('date_end'):
        try:
            plan_end = date.fromisoformat(plan['date_end'])
        except ValueError:
            plan_end = None
        if plan_end and plan_end > target_finish_date:
            errors.append(
                f'按当前方案预计到 {plan_end.isoformat()} 才能完成，晚于目标完成日 {target_finish_date.isoformat()}'
            )

    now = get_business_now()
    for session in session_dates:
        start_at = datetime.combine(
            date.fromisoformat(session['date']),
            datetime.strptime(session['time_start'], '%H:%M').time(),
        )
        if start_at < now:
            errors.append(f'{session["date"]} {session["time_start"]}-{session["time_end"]} 已是过去时间，不能保存')

    for current, nxt in zip(session_dates, session_dates[1:]):
        if current['date'] != nxt['date']:
            continue
        if _slots_overlap(
            current['time_start'],
            current['time_end'],
            nxt['time_start'],
            nxt['time_end'],
        ):
            errors.append(
                f'{current["date"]} 存在课次时间重叠：'
                f'{current["time_start"]}-{current["time_end"]} 与 {nxt["time_start"]}-{nxt["time_end"]}'
            )

    ignore_ids = []
    if enrollment and enrollment.id:
        ignore_ids = [schedule.id for schedule in _linked_schedule_query(enrollment.id).all()]

    errors.extend(
        _collect_teacher_schedule_conflicts(
            enrollment,
            session_dates,
            ignore_schedule_ids=ignore_ids,
        )
    )
    errors.extend(
        _collect_student_schedule_conflicts(
            enrollment,
            session_dates,
            ignore_schedule_ids=ignore_ids,
        )
    )

    teacher_context = _teacher_work_context(enrollment.teacher or enrollment.teacher_id)
    teacher_ranges = _load_teacher_available_ranges(enrollment.teacher_id)
    student_ranges = _student_available_slots_for_enrollment(enrollment)
    excluded_dates = _load_student_excluded_dates(enrollment.student_profile)
    intake_meta = _enrollment_availability_intake_meta(enrollment)

    for session in session_dates:
        if teacher_ranges and not _session_within_ranges(session, teacher_ranges):
            teacher_label = (
                '全职老师统一工作时段'
                if teacher_context['teacher_work_mode'] == TEACHER_WORK_MODE_FULL_TIME
                else '老师原始可用时间'
            )
            warnings.append(f'{session["date"]} {session["time_start"]}-{session["time_end"]} 超出{teacher_label}')
        if student_ranges and not _session_within_ranges(session, student_ranges):
            warnings.append(
                f'{session["date"]} {session["time_start"]}-{session["time_end"]} 超出学生填写的可上课时间'
            )
        if session['date'] in excluded_dates:
            warnings.append(f'{session["date"]} 命中学生标记的不可上课日期')
    if float(intake_meta.get('confidence') or 0) and float(intake_meta.get('confidence') or 0) < 0.75:
        warnings.append('学生时间解析置信度较低，建议让学生确认一次结构化结果')

    return list(dict.fromkeys(errors)), list(dict.fromkeys(warnings))


def _extract_legacy_enrollment_id(note_text):
    matches = {int(value) for value in LEGACY_ENROLLMENT_NOTE_PATTERN.findall(str(note_text or ''))}
    if len(matches) != 1:
        return None
    return next(iter(matches))


def _linked_schedule_ids(enrollment_id, *, include_cancelled=False):
    from modules.oa.models import CourseSchedule

    if not enrollment_id:
        return []

    direct_filters = [CourseSchedule.enrollment_id == enrollment_id]
    if not include_cancelled:
        direct_filters.append(CourseSchedule.is_cancelled == False)
    schedule_ids = {
        row[0]
        for row in db.session.query(CourseSchedule.id).filter(*direct_filters).all()
    }
    legacy_candidate_filters = [
        CourseSchedule.enrollment_id.is_(None),
        CourseSchedule.notes.isnot(None),
        CourseSchedule.notes.contains('报名#'),
    ]
    if not include_cancelled:
        legacy_candidate_filters.append(CourseSchedule.is_cancelled == False)
    legacy_candidates = CourseSchedule.query.filter(*legacy_candidate_filters).all()
    for schedule in legacy_candidates:
        if _extract_legacy_enrollment_id(schedule.notes) == enrollment_id:
            schedule_ids.add(schedule.id)
    return sorted(schedule_ids)


def _linked_schedule_query(enrollment_id, *, include_cancelled=False):
    from modules.oa.models import CourseSchedule

    linked_ids = _linked_schedule_ids(enrollment_id, include_cancelled=include_cancelled)
    if not linked_ids:
        return CourseSchedule.query.filter(False)
    return CourseSchedule.query.filter(CourseSchedule.id.in_(linked_ids))


def get_accessible_enrollment_query(user):
    from modules.auth.models import Enrollment

    query = Enrollment.query
    if not user or not getattr(user, 'is_authenticated', False):
        return query.filter(False)
    if user.role == 'admin':
        return query
    if user.role == 'teacher':
        return query.filter(Enrollment.teacher_id == user.id)
    if user.role == 'student':
        profile = user.student_profile
        if not profile:
            return query.filter(False)
        return query.filter(Enrollment.student_profile_id == profile.id)
    return query.filter(False)


def user_can_access_enrollment(user, enrollment):
    if not user or not enrollment or not getattr(user, 'is_authenticated', False):
        return False
    if user.role == 'admin':
        return True
    if user.role == 'teacher':
        return enrollment.teacher_id == user.id
    if user.role == 'student':
        profile = user.student_profile
        return bool(profile and enrollment.student_profile_id == profile.id)
    return False


def user_can_edit_enrollment_intake(user, enrollment):
    if not user_can_access_enrollment(user, enrollment):
        return False
    if user.role == 'teacher':
        return False
    return enrollment.status in {'pending_schedule', 'pending_student_confirm'}


def user_can_access_schedule(user, schedule):
    if not user or not schedule or not getattr(user, 'is_authenticated', False):
        return False
    if user.role == 'admin':
        return True
    if user.role == 'teacher':
        return _schedule_matches_teacher_actor(user, schedule)
    if user.role == 'student':
        profile = user.student_profile
        return bool(
            profile
            and _schedule_effective_student_profile_id(schedule) == profile.id
        )
    return False


def _latest_leave_request(schedule):
    from modules.auth.models import LeaveRequest

    if not schedule:
        return None
    return LeaveRequest.query.filter_by(schedule_id=schedule.id).order_by(
        LeaveRequest.created_at.desc()
    ).first()


def schedule_has_historical_facts(schedule, reference=None):
    from modules.auth.models import LeaveRequest

    if not schedule:
        return False
    if _schedule_has_started(schedule, reference):
        return True
    feedback = getattr(schedule, 'feedback', None)
    if feedback and feedback.status == 'submitted':
        return True
    return LeaveRequest.query.filter_by(schedule_id=schedule.id).count() > 0


def user_can_request_leave(user, schedule):
    latest_leave = _latest_leave_request(schedule)
    if not user_can_access_schedule(user, schedule):
        return False
    if user.role != 'student':
        return False
    if getattr(schedule, 'is_cancelled', False):
        return False
    if _schedule_has_started(schedule):
        return False
    if latest_leave and latest_leave.status in {'pending', 'approved'}:
        return False
    return True


def user_can_approve_leave(user, leave_request):
    if not user or not leave_request or not getattr(user, 'is_authenticated', False):
        return False
    if user.role == 'admin':
        return True
    if user.role != 'teacher':
        return False
    return bool(leave_request.schedule and _schedule_matches_teacher_actor(user, leave_request.schedule))


def user_can_chat_with(user, partner):
    from modules.auth.models import Enrollment

    if not user or not partner or not getattr(user, 'is_authenticated', False):
        return False
    if not getattr(partner, 'is_active', False):
        return False
    if user.id == partner.id:
        return False
    if user.role == 'admin' or partner.role == 'admin':
        return True
    if user.role == 'teacher' and partner.role == 'student':
        partner_profile = partner.student_profile
        if not partner_profile:
            return False
        return Enrollment.query.filter_by(
            teacher_id=user.id,
            student_profile_id=partner_profile.id,
        ).count() > 0
    if user.role == 'student' and partner.role == 'teacher':
        profile = user.student_profile
        if not profile:
            return False
        return Enrollment.query.filter_by(
            teacher_id=partner.id,
            student_profile_id=profile.id,
        ).count() > 0
    return False


def user_can_access_chat_history(user, partner):
    from modules.auth.models import ChatMessage

    if not user or not partner or not getattr(user, 'is_authenticated', False):
        return False
    if user.id == getattr(partner, 'id', None):
        return False
    if user_can_chat_with(user, partner):
        return True

    return ChatMessage.query.filter(
        or_(
            and_(ChatMessage.sender_id == user.id, ChatMessage.receiver_id == partner.id),
            and_(ChatMessage.sender_id == partner.id, ChatMessage.receiver_id == user.id),
        )
    ).count() > 0


def user_can_submit_feedback(user, schedule):
    if not (
        user
        and getattr(user, 'is_authenticated', False)
        and schedule
        and _actor_can_manage_schedule_feedback(user, schedule)
    ):
        return False
    if not schedule_requires_course_feedback(schedule):
        return False
    if not _schedule_has_started(schedule):
        return False
    latest_leave = _latest_leave_request(schedule)
    if latest_leave and latest_leave.status == 'approved':
        return False
    feedback = getattr(schedule, 'feedback', None)
    if feedback and feedback.status == 'submitted':
        return False
    return True


def get_schedule_feedback_permission_error(user, schedule):
    if not schedule or not _actor_can_manage_schedule_feedback(user, schedule):
        return '无权提交该课程反馈'
    if not schedule_requires_course_feedback(schedule):
        return get_course_feedback_skip_reason(schedule) or '该课程当前无需填写反馈'
    feedback = getattr(schedule, 'feedback', None)
    if feedback and feedback.status == 'submitted':
        return '该课程反馈已提交'
    latest_leave = _latest_leave_request(schedule)
    if latest_leave and latest_leave.status == 'approved':
        return '该课程已批准请假，不能提交反馈'
    if not _schedule_has_started(schedule):
        return '课程尚未开始，暂不能提交反馈'
    return None


def build_feedback_payload(feedback, actor=None):
    if not feedback:
        return None
    payload = feedback.to_dict()
    payload['can_submit_feedback'] = bool(
        actor
        and getattr(actor, 'is_authenticated', False)
        and _actor_can_manage_schedule_feedback(actor, teacher_id=feedback.teacher_id)
    )
    return payload


def _datetime_to_iso(value):
    return value.isoformat() if value else None


def _parse_iso_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None


def build_next_action_meta(*, role=None, label=None, status=None, waiting_since=None, is_overdue=False):
    return {
        'next_action_role': role,
        'next_action_label': label,
        'next_action_status': status,
        'waiting_since': _datetime_to_iso(waiting_since),
        'is_overdue': bool(is_overdue),
    }


def get_workflow_next_action_meta(todo, *, payload=None):
    from modules.oa.models import OATodo

    if not todo:
        return build_next_action_meta()

    workflow_payload = payload or todo.get_payload_data()
    role = None
    label = None
    status = todo.workflow_status
    waiting_since = todo.updated_at or todo.created_at

    if todo.workflow_status == OATodo.WORKFLOW_STATUS_WAITING_TEACHER_PROPOSAL:
        role = 'teacher'
        if todo.todo_type == OATodo.TODO_TYPE_SCHEDULE_FEEDBACK:
            label = '提交课后反馈'
            status = 'waiting_teacher_feedback'
            if todo.schedule:
                waiting_since = _schedule_start_datetime(todo.schedule)
        elif todo.todo_type == OATodo.TODO_TYPE_LEAVE_MAKEUP:
            label = '提交补课建议'
        else:
            label = '提交排课建议'
        latest_rejection = workflow_payload.get('latest_rejection') or {}
        waiting_since = _parse_iso_datetime(latest_rejection.get('created_at')) or waiting_since
    elif todo.workflow_status == OATodo.WORKFLOW_STATUS_WAITING_ADMIN_REVIEW:
        role = 'admin'
        label = '微调并发送给学生'
        proposal = workflow_payload.get('current_proposal') or {}
        waiting_since = _parse_iso_datetime(proposal.get('submitted_at')) or waiting_since
    elif todo.workflow_status == OATodo.WORKFLOW_STATUS_WAITING_STUDENT_CONFIRM:
        role = 'student'
        label = '确认补课方案' if todo.todo_type == OATodo.TODO_TYPE_LEAVE_MAKEUP else '确认排课方案'
        waiting_since = _parse_iso_datetime(workflow_payload.get('sent_to_student_at')) or waiting_since
    elif todo.workflow_status == OATodo.WORKFLOW_STATUS_COMPLETED:
        role = 'none'
        label = '已完成'
        waiting_since = todo.completed_at or waiting_since
    elif todo.workflow_status == OATodo.WORKFLOW_STATUS_CANCELLED:
        role = 'none'
        label = '已取消'
        waiting_since = todo.completed_at or waiting_since

    is_overdue = bool(
        todo.is_open_workflow
        and todo.due_date
        and todo.due_date < get_business_today()
    )
    return build_next_action_meta(
        role=role,
        label=label,
        status=status,
        waiting_since=waiting_since,
        is_overdue=is_overdue,
    )


def _filter_workflow_todos_for_actor(workflow_todos, actor=None):
    if not (actor and getattr(actor, 'is_authenticated', False) and actor.role == 'student'):
        return workflow_todos
    visible_todo_types = {'enrollment_replan', 'leave_makeup'}
    return [
        todo for todo in (workflow_todos or [])
        if todo.get('todo_type') in visible_todo_types
    ]


def build_schedule_payload(schedule, actor=None):
    from modules.auth.workflow_services import get_schedule_workflow_todos
    from modules.oa.services import delivery_mode_label, meeting_status_label

    payload = schedule.to_dict()
    latest_leave = _latest_leave_request(schedule)
    feedback = getattr(schedule, 'feedback', None)
    feedback_required = schedule_requires_course_feedback(schedule)
    workflow_todos = get_schedule_workflow_todos(schedule.id, actor)
    if actor and getattr(actor, 'is_authenticated', False) and actor.role == 'student':
        if feedback and feedback.status != 'submitted':
            feedback = None
        workflow_todos = _filter_workflow_todos_for_actor(workflow_todos, actor)
    payload.update({
        'delivery_mode_label': delivery_mode_label(payload.get('delivery_mode')),
        'meeting_status_label': meeting_status_label(payload.get('meeting_status')),
        'meeting_material': schedule.meeting_material.to_dict() if getattr(schedule, 'meeting_material', None) else None,
        'leave_request': latest_leave.to_dict() if latest_leave else None,
        'leave_status': latest_leave.status if latest_leave else None,
        'feedback': build_feedback_payload(feedback, actor),
        'feedback_status': feedback.status if feedback else None,
        'feedback_submitted_at': feedback.submitted_at.isoformat() if feedback and feedback.submitted_at else None,
        'is_delivered': bool(feedback and feedback.status == 'submitted'),
        'can_edit': bool(
            actor
            and getattr(actor, 'is_authenticated', False)
            and actor.role == 'admin'
            and not payload.get('is_cancelled')
        ),
        'can_confirm': False,
        'can_reject': False,
        'can_request_leave': bool(actor and user_can_request_leave(actor, schedule)),
        'can_approve_leave': bool(actor and latest_leave and latest_leave.status == 'pending' and user_can_approve_leave(actor, latest_leave)),
        'can_submit_feedback': bool(actor and user_can_submit_feedback(actor, schedule)),
        'workflow_todos': workflow_todos,
        'feedback_due_at': (
            _datetime_to_iso(_schedule_end_datetime(schedule))
            if feedback_required and _schedule_has_started(schedule)
            else None
        ),
        'feedback_delay_days': (
            max((get_business_today() - schedule.date).days, 0)
            if feedback_required and _schedule_has_started(schedule)
            else 0
        ),
        'missing_feedback_count_for_teacher_recent': 0,
        'is_repeat_late_teacher': False,
    })
    if payload.get('is_cancelled'):
        payload.update(build_next_action_meta(
            role='none',
            label='课次已取消',
            status='cancelled',
            waiting_since=payload.get('cancelled_at') or payload.get('updated_at') or payload.get('created_at'),
            is_overdue=False,
        ))
        return payload
    active_todo = next(
        (
            item for item in workflow_todos
            if item.get('workflow_status') not in {'completed', 'cancelled'}
        ),
        None,
    )
    if active_todo:
        payload.update({
            'next_action_role': active_todo.get('next_action_role'),
            'next_action_label': active_todo.get('next_action_label'),
            'next_action_status': active_todo.get('next_action_status'),
            'waiting_since': active_todo.get('waiting_since'),
            'is_overdue': bool(active_todo.get('is_overdue')),
        })
    elif feedback_required and (
        user_can_submit_feedback(actor, schedule) or (
            _schedule_has_started(schedule)
            and not (feedback and feedback.status == 'submitted')
            and not (latest_leave and latest_leave.status == 'approved')
        )
    ):
        payload.update(build_next_action_meta(
            role='teacher',
            label='提交课后反馈',
            status='waiting_teacher_feedback',
            waiting_since=_schedule_start_datetime(schedule),
            is_overdue=bool(schedule.date < get_business_today()),
        ))
    else:
        payload.update(build_next_action_meta(
            role='none',
            label='等待课程进行' if _schedule_start_datetime(schedule) > get_business_now() else '已完成当前动作',
            status='idle',
            waiting_since=schedule.updated_at or schedule.created_at,
            is_overdue=False,
        ))
    return payload


def build_leave_request_payload(leave_request, actor=None):
    from modules.auth.workflow_services import get_leave_request_workflow

    payload = leave_request.to_dict()
    makeup_workflow = get_leave_request_workflow(leave_request.id, actor)
    makeup_schedule = build_schedule_payload(leave_request.makeup_schedule, actor) if leave_request.makeup_schedule else None
    makeup_preference_summary = _summarize_makeup_preferences(
        payload.get('makeup_available_slots'),
        payload.get('makeup_excluded_dates'),
        payload.get('makeup_preference_note'),
    )
    makeup_status = (
        'confirmed'
        if leave_request.makeup_schedule_id
        else 'waiting_arrangement'
        if leave_request.status == 'approved'
        else None
    )
    payload.update({
        'can_edit': False,
        'can_confirm': False,
        'can_reject': False,
        'can_request_leave': False,
        'can_approve_leave': bool(
            actor
            and leave_request.status == 'pending'
            and user_can_approve_leave(actor, leave_request)
        ),
        'can_submit_feedback': False,
        'makeup_workflow': makeup_workflow,
        'makeup_schedule_id': leave_request.makeup_schedule_id,
        'makeup_schedule': makeup_schedule,
        'makeup_status': makeup_status,
        'original_schedule_summary': _summarize_schedule_summary(leave_request.schedule),
        'makeup_preference_summary': makeup_preference_summary,
        'decision_comment': payload.get('decision_comment'),
        'context_summary': makeup_preference_summary,
        'latest_rejection_text': None,
        'proposal_note': None,
        'current_plan_summary': _summarize_schedule_summary(leave_request.makeup_schedule) if leave_request.makeup_schedule else None,
        'case_stage': None,
        'case_stage_label': None,
        'related_workflow_id': makeup_workflow.get('id') if makeup_workflow else None,
    })
    if makeup_workflow:
        payload.update({
            'next_action_role': makeup_workflow.get('next_action_role'),
            'next_action_label': makeup_workflow.get('next_action_label'),
            'next_action_status': makeup_workflow.get('next_action_status'),
            'waiting_since': makeup_workflow.get('waiting_since'),
            'is_overdue': bool(makeup_workflow.get('is_overdue')),
            'context_summary': makeup_workflow.get('context_summary') or payload.get('context_summary'),
            'latest_rejection_text': makeup_workflow.get('latest_rejection_text'),
            'proposal_note': makeup_workflow.get('proposal_note'),
            'current_plan_summary': makeup_workflow.get('current_plan_summary') or payload.get('current_plan_summary'),
            'case_stage': makeup_workflow.get('next_action_status'),
        })
    elif leave_request.status == 'pending':
        payload.update(build_next_action_meta(
            role='teacher_or_admin',
            label='审批请假申请',
            status='waiting_leave_approval',
            waiting_since=leave_request.created_at,
            is_overdue=False,
        ))
        payload['case_stage'] = 'waiting_leave_approval'
    elif leave_request.status == 'approved' and leave_request.makeup_schedule_id:
        payload.update(build_next_action_meta(
            role='none',
            label='补课已确认',
            status='confirmed',
            waiting_since=leave_request.makeup_schedule.updated_at if leave_request.makeup_schedule else leave_request.created_at,
            is_overdue=False,
        ))
        payload['case_stage'] = 'confirmed'
    elif leave_request.status == 'approved':
        payload.update(build_next_action_meta(
            role='teacher',
            label='提交补课建议',
            status='waiting_teacher_proposal',
            waiting_since=leave_request.created_at,
            is_overdue=False,
        ))
        payload['case_stage'] = 'waiting_teacher_proposal'
    else:
        payload.update(build_next_action_meta(
            role='none',
            label='请假流程已结束',
            status=leave_request.status,
            waiting_since=leave_request.created_at,
            is_overdue=False,
        ))
        payload['case_stage'] = leave_request.status
    payload['case_stage_label'] = {
        'waiting_leave_approval': '待审批',
        'waiting_teacher_proposal': '待老师提案',
        'waiting_admin_review': '待教务发送',
        'waiting_student_confirm': '待学生确认补课',
        'confirmed': '已确认补课',
        'cancelled': '已取消',
        'rejected': '已驳回',
        'approved': '补课安排中',
    }.get(payload.get('case_stage'), payload.get('next_action_label') or payload.get('status'))
    payload['next_step_hint'] = {
        'waiting_leave_approval': '审批通过后会进入补课安排。',
        'waiting_teacher_proposal': '老师提交补课建议后，教务会继续发送给学生确认。',
        'waiting_admin_review': '教务确认后会把补课方案发给学生。',
        'waiting_student_confirm': '学生确认后补课时间会正式生效。',
        'confirmed': '请按已确认的补课时间正常上课。',
        'cancelled': '该请假关联课次已取消，请按最新安排与教务沟通。',
        'rejected': '请查看处理说明后重新发起请假。',
    }.get(payload.get('case_stage'))
    return payload


def _get_enrollment_delivery_meta(enrollment):
    from modules.oa.models import CourseFeedback

    schedules = _linked_schedule_query(enrollment.id).all()
    schedule_ids = [schedule.id for schedule in schedules]
    latest_feedback = None
    completed_count = 0
    pending_feedback_count = 0
    approved_leave_count = 0

    if schedule_ids:
        submitted_feedbacks = CourseFeedback.query.filter(
            CourseFeedback.schedule_id.in_(schedule_ids),
            CourseFeedback.status == 'submitted',
        ).order_by(CourseFeedback.submitted_at.desc(), CourseFeedback.updated_at.desc()).all()
        completed_count = len(submitted_feedbacks)
        latest_feedback = submitted_feedbacks[0] if submitted_feedbacks else None

    now = get_business_now()
    for schedule in schedules:
        latest_leave = _latest_leave_request(schedule)
        if latest_leave and latest_leave.status == 'approved':
            approved_leave_count += 1
        feedback = getattr(schedule, 'feedback', None)
        if (
            schedule_requires_course_feedback(schedule)
            and _schedule_has_started(schedule, now)
            and not (feedback and feedback.status == 'submitted')
        ):
            if not latest_leave or latest_leave.status != 'approved':
                pending_feedback_count += 1

    return {
        'scheduled_count': len(schedules),
        'completed_count': completed_count,
        'leave_count': approved_leave_count,
        'pending_feedback_count': pending_feedback_count,
        'latest_teacher_feedback': latest_feedback.summary if latest_feedback and latest_feedback.summary else None,
        'latest_teacher_feedback_at': latest_feedback.submitted_at.isoformat() if latest_feedback and latest_feedback.submitted_at else None,
        'latest_teacher_feedback_detail': build_feedback_payload(latest_feedback),
    }


def get_enrollment_feedback_meta(enrollment):
    """返回报名最近一次学生排课反馈及未读状态，优先读取 workflow 结构化 rejection。"""
    from modules.auth.models import ChatMessage
    from modules.oa.models import OATodo

    if not enrollment:
        return {
            'latest_feedback': None,
            'latest_feedback_at': None,
            'has_unread_feedback': False,
        }

    latest_workflow = OATodo.query.filter(
        OATodo.enrollment_id == enrollment.id,
        OATodo.todo_type == OATodo.TODO_TYPE_ENROLLMENT_REPLAN,
    ).order_by(OATodo.updated_at.desc(), OATodo.created_at.desc()).first()
    if latest_workflow:
        workflow_payload = latest_workflow.get_payload_data()
        latest_rejection = workflow_payload.get('latest_rejection') or {}
        feedback_text = (latest_rejection.get('message') or latest_rejection.get('reason') or '').strip()
        if feedback_text:
            rejections = workflow_payload.get('rejections') or []
            unread_reference = latest_rejection.get('created_at')
            has_unread = False
            if unread_reference:
                latest_chat = ChatMessage.query.filter(
                    ChatMessage.enrollment_id == enrollment.id,
                    ChatMessage.content.startswith(FEEDBACK_PREFIX),
                    ChatMessage.is_read == False,
                ).order_by(ChatMessage.created_at.desc()).first()
                if latest_chat and latest_chat.created_at:
                    has_unread = latest_chat.created_at.isoformat() <= unread_reference
            return {
                'latest_feedback': feedback_text,
                'latest_feedback_at': latest_rejection.get('created_at'),
                'has_unread_feedback': has_unread or bool(rejections and latest_workflow.is_open_workflow),
            }

    query = ChatMessage.query.filter(
        ChatMessage.enrollment_id == enrollment.id,
        ChatMessage.content.startswith(FEEDBACK_PREFIX),
    )
    latest = query.order_by(ChatMessage.created_at.desc()).first()
    if not latest:
        return {
            'latest_feedback': None,
            'latest_feedback_at': None,
            'has_unread_feedback': False,
        }

    feedback_text = latest.content[len(FEEDBACK_PREFIX):].strip()
    if feedback_text.startswith('[') and '] ' in feedback_text:
        feedback_text = feedback_text.split('] ', 1)[1]
    has_unread = query.filter(ChatMessage.is_read == False).count() > 0
    return {
        'latest_feedback': feedback_text,
        'latest_feedback_at': latest.created_at.isoformat() if latest.created_at else None,
        'has_unread_feedback': has_unread,
    }


def build_enrollment_payload(enrollment, actor=None):
    from modules.auth.workflow_services import get_enrollment_workflow_todos
    from modules.oa.services import delivery_mode_label

    payload = enrollment.to_dict()
    profile_meta = _profile_constraint_meta(enrollment.student_profile)
    availability_meta = payload.get('availability_intake') or _enrollment_availability_intake_meta(enrollment)
    recommended_bundle = normalize_plan(payload.get('recommended_bundle'), enrollment)
    risk_assessment = payload.get('risk_assessment') or {}
    teacher_context = _teacher_work_context(enrollment.teacher or enrollment.teacher_id)
    payload['proposed_slots'] = [
        normalize_plan(plan, enrollment)
        for plan in payload.get('proposed_slots', [])
    ]
    payload['confirmed_slot'] = normalize_plan(payload.get('confirmed_slot'), enrollment)
    payload['recommended_bundle'] = recommended_bundle
    payload.update(get_enrollment_feedback_meta(enrollment))
    payload.update(_get_enrollment_delivery_meta(enrollment))
    payload.update({
        **teacher_context,
        'delivery_preference_label': delivery_mode_label(payload.get('delivery_preference')),
        'delivery_urgency_label': DELIVERY_URGENCY_LABELS.get(payload.get('delivery_urgency'), payload.get('delivery_urgency') or '常规'),
        'expected_session_count': _get_total_sessions(enrollment),
        'can_edit': bool(actor and user_can_edit_enrollment_intake(actor, enrollment)),
        'can_confirm': bool(
            actor
            and getattr(actor, 'is_authenticated', False)
            and actor.role == 'student'
            and user_can_access_enrollment(actor, enrollment)
            and enrollment.status == 'pending_student_confirm'
        ),
        'can_reject': bool(
            actor
            and getattr(actor, 'is_authenticated', False)
            and actor.role == 'student'
            and user_can_access_enrollment(actor, enrollment)
            and enrollment.status == 'pending_student_confirm'
        ),
        'can_request_leave': False,
        'can_approve_leave': False,
        'can_submit_feedback': False,
        'edit_intake_url': f'/auth/enrollments/{enrollment.id}/intake-edit',
        'student_availability_summary': profile_meta['available_times_summary'],
        'excluded_dates_summary': profile_meta['excluded_dates_summary'],
        'availability_intake_summary': availability_meta.get('summary'),
        'availability_confidence': availability_meta.get('confidence'),
        'availability_needs_review': bool(availability_meta.get('needs_review')),
        'candidate_slot_pool': payload.get('candidate_slot_pool') or [],
        'risk_assessment': risk_assessment,
        'scheduling_complexity_hint': _build_scheduling_complexity_hint(enrollment),
    })
    workflow_todos = _filter_workflow_todos_for_actor(
        get_enrollment_workflow_todos(enrollment.id, actor),
        actor,
    )
    payload['workflow_todos'] = workflow_todos
    payload['active_workflow_todo'] = workflow_todos[0] if workflow_todos else None
    payload['current_plan_summary'] = _summarize_plan(payload.get('confirmed_slot'))
    payload['proposal_note'] = ((payload.get('confirmed_slot') or {}).get('note') or '').strip() or None
    payload['latest_rejection_text'] = payload.get('latest_feedback')
    payload['context_summary'] = payload.get('next_action_label')
    payload['current_plan_session_dates'] = ((payload.get('confirmed_slot') or {}).get('session_dates') or [])
    payload['session_preview_lines'] = _session_preview_lines(payload.get('current_plan_session_dates'))
    active_todo = payload['active_workflow_todo']
    if active_todo:
        payload.update({
            'next_action_role': active_todo.get('next_action_role'),
            'next_action_label': active_todo.get('next_action_label'),
            'next_action_status': active_todo.get('next_action_status'),
            'waiting_since': active_todo.get('waiting_since'),
            'is_overdue': bool(active_todo.get('is_overdue')),
            'context_summary': active_todo.get('context_summary') or payload.get('context_summary'),
            'latest_rejection_text': active_todo.get('latest_rejection_text') or payload.get('latest_rejection_text'),
            'proposal_note': active_todo.get('proposal_note') or payload.get('proposal_note'),
            'current_plan_summary': active_todo.get('current_plan_summary') or payload.get('current_plan_summary'),
            'original_schedule_summary': active_todo.get('original_schedule_summary'),
            'current_plan_session_dates': (
                ((active_todo.get('current_proposal') or {}).get('session_dates'))
                or active_todo.get('current_plan_session_dates')
                or payload.get('current_plan_session_dates')
            ),
        })
        payload['session_preview_lines'] = _session_preview_lines(payload.get('current_plan_session_dates'))
    elif enrollment.status == 'pending_info':
        payload.update(build_next_action_meta(
            role='student',
            label='提交报名信息',
            status='waiting_student_info',
            waiting_since=enrollment.created_at,
            is_overdue=False,
        ))
    elif enrollment.status == 'pending_schedule':
        pending_schedule_label = '安排初始排课'
        pending_schedule_role = 'admin'
        pending_schedule_status = 'waiting_admin_schedule'
        if (risk_assessment.get('recommended_action') or '') == 'needs_student_clarification':
            pending_schedule_label = '补充可上课时间'
            pending_schedule_role = 'student'
            pending_schedule_status = 'waiting_student_clarification'
        elif (risk_assessment.get('recommended_action') or '') == 'needs_admin_intervention':
            pending_schedule_label = '补齐排课风险'
        elif (risk_assessment.get('recommended_action') or '') == 'needs_teacher_confirmation':
            pending_schedule_label = '联系老师确认模板外时段'
        elif not payload.get('recommended_bundle'):
            pending_schedule_label = '补充可上课时间'
        payload.update(build_next_action_meta(
            role=pending_schedule_role,
            label=pending_schedule_label,
            status=pending_schedule_status,
            waiting_since=enrollment.updated_at or enrollment.created_at,
            is_overdue=False,
        ))
    elif enrollment.status == 'pending_student_confirm':
        payload.update(build_next_action_meta(
            role='student',
            label='确认排课方案',
            status='waiting_student_confirm',
            waiting_since=enrollment.updated_at or enrollment.created_at,
            is_overdue=False,
        ))
    elif payload.get('pending_feedback_count'):
        payload.update(build_next_action_meta(
            role='teacher',
            label='提交课后反馈',
            status='waiting_teacher_feedback',
            waiting_since=enrollment.updated_at or enrollment.created_at,
            is_overdue=bool(payload.get('pending_feedback_count')),
        ))
    else:
        label = '交付进行中'
        if enrollment.status == 'completed':
            label = '已完成'
        elif enrollment.status == 'confirmed':
            label = '等待首次上课'
        payload.update(build_next_action_meta(
            role='none',
            label=label,
            status=enrollment.status,
            waiting_since=enrollment.updated_at or enrollment.created_at,
            is_overdue=False,
        ))
    if not payload.get('context_summary'):
        payload['context_summary'] = payload.get('next_action_label')
    payload['current_stage_label'] = {
        'waiting_student_clarification': '待学生补充时间',
        'waiting_student_confirm': '待学生确认方案',
        'waiting_admin_schedule': '待教务排课',
        'waiting_teacher_feedback': '待老师提交反馈',
        'confirmed': '等待首次上课',
        'completed': '已完成',
    }.get(payload.get('next_action_status'), payload.get('next_action_label'))
    payload['next_step_hint'] = {
        'waiting_student_clarification': '请补充或确认可上课时间，系统再继续排课。',
        'waiting_student_confirm': '学生确认后系统会生成正式课表。',
        'waiting_admin_schedule': (
            '当前方案存在风险，教务需要先补齐时间或调整方案。'
            if (risk_assessment.get('recommended_action') or '') == 'needs_admin_intervention'
            else (
                '当前方案需要老师确认模板外时段，教务需先发起老师协同。'
                if (risk_assessment.get('recommended_action') or '') == 'needs_teacher_confirmation'
                else '教务排好课后会发送给学生确认。'
            )
        ),
        'waiting_teacher_feedback': '老师提交反馈后，学生和教务都可查看。',
    }.get(payload.get('next_action_status'))
    return payload


def sync_enrollment_status(enrollment):
    """根据课表与交付情况重新计算报名状态。"""
    from modules.oa.models import CourseFeedback, CourseSchedule

    if not enrollment:
        return None
    if enrollment.status in {'pending_info', 'pending_schedule'}:
        return enrollment.status

    schedules = _linked_schedule_query(enrollment.id).order_by(
        CourseSchedule.date,
        CourseSchedule.time_start,
    ).all()
    if enrollment.status == 'pending_student_confirm' and not schedules:
        return enrollment.status
    if not schedules:
        enrollment.status = 'confirmed' if enrollment.confirmed_slot else 'pending_schedule'
        return enrollment.status

    now = get_business_now()
    started_schedules = [schedule for schedule in schedules if _schedule_has_started(schedule, now)]
    if not started_schedules:
        enrollment.status = 'confirmed'
        return enrollment.status

    has_future_schedule = any(_schedule_start_datetime(schedule) > now for schedule in schedules)
    submitted_feedback_schedule_ids = set()
    if started_schedules:
        submitted_feedback_schedule_ids = {
            row[0]
            for row in db.session.query(CourseFeedback.schedule_id).filter(
                CourseFeedback.schedule_id.in_([schedule.id for schedule in started_schedules]),
                CourseFeedback.status == 'submitted',
            ).all()
        }
    has_open_delivery = False
    for schedule in started_schedules:
        if not schedule_requires_course_feedback(schedule):
            continue
        latest_leave = _latest_leave_request(schedule)
        approved_leave = bool(latest_leave and latest_leave.status == 'approved')
        delivered = schedule.id in submitted_feedback_schedule_ids
        if not delivered and not approved_leave:
            has_open_delivery = True
            break

    enrollment.status = 'active' if has_future_schedule or has_open_delivery else 'completed'
    return enrollment.status


def find_matching_slots(enrollment_id):
    """自动匹配排课，按 sessions_per_week 生成推荐方案。"""
    from modules.auth.models import Enrollment
    from modules.auth.workflow_services import ensure_enrollment_replan_workflow

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return [], '报名记录不存在'
    if not enrollment.student_profile:
        return [], '学生尚未提交可上课时间'

    state = refresh_enrollment_scheduling_ai_state(enrollment)
    risk_assessment = state.get('risk_assessment') or {}
    hard_errors = risk_assessment.get('hard_errors') or []
    if hard_errors:
        return [], hard_errors[0]
    if (risk_assessment.get('recommended_action') or '') == 'needs_student_clarification':
        clarification_reasons = risk_assessment.get('clarification_reasons') or []
        return [], clarification_reasons[0] if clarification_reasons else (
            risk_assessment.get('summary') or '当前还需要学生补充可上课时间'
        )
    if (risk_assessment.get('recommended_action') or '') == 'needs_teacher_confirmation':
        ensure_enrollment_replan_workflow(
            enrollment,
            rejection_text='AI 判断当前需老师确认工作模板外时段，请老师补充可执行方案。',
            actor_user=None,
        )
        db.session.commit()
        return [], risk_assessment.get('summary') or '当前需老师确认模板外时段，请先发起老师协同'

    proposed_plans = state.get('proposed_plans') or []
    enrollment.proposed_slots = json.dumps(proposed_plans, ensure_ascii=False) if proposed_plans else None
    db.session.flush()
    if not proposed_plans:
        return [], '当前没有可发送给学生的稳定方案'
    return proposed_plans, None


def propose_enrollment_schedule(enrollment_id, slot_index):
    """管理员选择 plan -> 保存并通知学生确认。"""
    from modules.auth.models import Enrollment
    from modules.auth.workflow_services import has_open_enrollment_replan_workflow

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return False, '报名记录不存在', []
    if enrollment.status not in {'pending_schedule', 'pending_student_confirm'}:
        return False, '当前状态不允许重新发送排课方案', []
    if has_open_enrollment_replan_workflow(enrollment.id):
        return False, '当前报名存在进行中的排课工作流，请通过工作流继续处理', []
    if enrollment.status == 'pending_student_confirm' and _linked_schedule_ids(enrollment.id):
        return False, '当前报名已存在正式课表，不能重新发送排课方案', []

    if not enrollment.proposed_slots:
        return False, '没有可用的排课方案', []

    try:
        proposed = json.loads(enrollment.proposed_slots)
    except (json.JSONDecodeError, TypeError):
        return False, '排课方案数据异常', []

    if slot_index < 0 or slot_index >= len(proposed):
        return False, '方案索引无效', []

    raw_plan = proposed[slot_index]
    plan = normalize_plan(raw_plan, enrollment)
    if not plan or not plan.get('weekly_slots'):
        return False, '排课方案数据异常', []

    enrollment.confirmed_slot = json.dumps(plan, ensure_ascii=False)
    enrollment.status = 'pending_student_confirm'
    _send_schedule_notification(enrollment, plan)
    db.session.commit()

    msg = f'已通知学生确认，共 {len(plan["session_dates"])} 节课'
    if plan['skipped_dates']:
        msg += f'（跳过 {len(plan["skipped_dates"])} 个不可上课日期）'
    return True, msg, plan['session_dates']


def save_manual_enrollment_plan(enrollment_id, session_dates, *, force_save=False):
    from modules.auth.models import Enrollment
    from modules.auth.workflow_services import has_open_enrollment_replan_workflow

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return {
            'success': False,
            'status_code': 404,
            'error': '报名记录不存在',
        }

    if enrollment.status not in {'pending_schedule', 'pending_student_confirm'}:
        return {
            'success': False,
            'status_code': 400,
            'error': '当前状态不允许手动调整排课方案',
        }
    if has_open_enrollment_replan_workflow(enrollment.id):
        return {
            'success': False,
            'status_code': 400,
            'error': '当前报名存在进行中的排课工作流，请通过工作流继续处理',
        }

    normalized_dates, parse_errors = _normalize_manual_session_dates(session_dates)
    if parse_errors:
        return {
            'success': False,
            'status_code': 400,
            'error': '；'.join(parse_errors),
            'errors': parse_errors,
        }

    errors, warnings = _collect_manual_plan_issues(enrollment, normalized_dates)
    if errors:
        return {
            'success': False,
            'status_code': 400,
            'error': '；'.join(errors),
            'errors': errors,
        }

    if warnings and not force_save:
        return {
            'success': False,
            'status_code': 200,
            'error': '方案存在提示项，请确认是否继续保存',
            'warnings': warnings,
            'can_force_save': True,
        }

    plan = _build_manual_plan(normalized_dates)
    if warnings:
        plan['manual_warnings'] = warnings
    enrollment.confirmed_slot = json.dumps(plan, ensure_ascii=False)
    enrollment.status = 'pending_student_confirm'
    _send_schedule_notification(enrollment, plan)
    db.session.commit()

    return {
        'success': True,
        'status_code': 200,
        'message': '手动排课方案已保存，并已重新通知学生确认',
        'plan': plan,
    }


def _send_schedule_notification(enrollment, plan):
    """通过 ChatMessage 通知学生查看排课方案。"""
    from modules.auth.models import ChatMessage

    profile = enrollment.student_profile
    if not profile or not profile.user_id:
        return

    day_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    weekly_text = '；'.join(
        f'{day_names[slot["day_of_week"]]} {slot["time_start"]}-{slot["time_end"]}'
        for slot in plan.get('weekly_slots', [])
    )
    first_date = plan.get('date_start') or '待定'
    last_date = plan.get('date_end') or '待定'

    content = (
        f'{SCHEDULE_PREFIX}[{enrollment.course_name}] '
        f'你的课程排课方案已准备好：\n'
        f'每周安排：{weekly_text or "待定"}\n'
        f'共 {len(plan.get("session_dates", []))} 节，预计 {plan.get("estimated_weeks", 0)} 周完成\n'
        f'首次上课：{first_date}\n'
        f'最后一课：{last_date}\n'
        '请前往「我的课表」页面查看并确认。'
    )

    msg = ChatMessage(
        sender_id=enrollment.teacher_id,
        receiver_id=profile.user_id,
        enrollment_id=enrollment.id,
        content=content,
        is_read=False,
    )
    db.session.add(msg)


def student_confirm_schedule(enrollment_id):
    """学生确认排课 -> 创建全部具体课次。"""
    from modules.auth.models import Enrollment
    from modules.oa.models import CourseFeedback, CourseSchedule, OATodo
    from modules.oa.services import build_schedule_delivery_fields
    from modules.auth.models import LeaveRequest
    from modules.auth.workflow_services import (
        complete_replan_workflows_for_enrollment,
        ensure_enrollment_replan_workflow,
        ensure_schedule_feedback_todo,
    )

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return False, '报名记录不存在', 0

    if enrollment.status != 'pending_student_confirm':
        return False, '当前状态不允许确认', 0

    def _reopen_replan(message_text):
        ensure_enrollment_replan_workflow(
            enrollment,
            rejection_text=message_text,
            actor_user=None,
        )
        enrollment.confirmed_slot = None
        enrollment.status = 'pending_schedule'
        db.session.commit()

    if not enrollment.confirmed_slot:
        message = '当前排课方案已失效，请等待老师和教务重新排课'
        _reopen_replan(message)
        return False, message, 0

    try:
        plan = normalize_plan(json.loads(enrollment.confirmed_slot), enrollment)
    except (json.JSONDecodeError, TypeError):
        message = '当前排课方案已失效，请等待老师和教务重新排课'
        _reopen_replan(message)
        return False, message, 0

    session_dates = plan.get('session_dates', [])
    if not session_dates:
        message = '当前排课方案已失效，请等待老师和教务重新排课'
        _reopen_replan(message)
        return False, message, 0

    existing_ids = [schedule.id for schedule in _linked_schedule_query(enrollment.id).all()]
    if existing_ids:
        return False, '当前报名已存在正式课表，不能重复确认排课方案', 0

    errors, _ = _collect_manual_plan_issues(
        enrollment,
        session_dates,
        weekly_slots=plan.get('weekly_slots') if isinstance(plan, dict) else None,
    )
    if errors:
        message = '；'.join(dict.fromkeys(errors))
        _reopen_replan(message)
        return False, message, 0

    teacher_name = enrollment.teacher.display_name if enrollment.teacher else ''
    student_name = enrollment.student_name

    created_count = 0
    for session in session_dates:
        course_date = date.fromisoformat(session['date'])
        delivery_fields = build_schedule_delivery_fields(
            delivery_mode=enrollment.delivery_preference,
            fallback_delivery_mode=enrollment.delivery_preference,
            allow_unknown=False,
        )
        schedule = CourseSchedule(
            date=course_date,
            day_of_week=session['day_of_week'],
            time_start=session['time_start'],
            time_end=session['time_end'],
            teacher=teacher_name,
            teacher_id=enrollment.teacher_id,
            course_name=enrollment.course_name,
            enrollment_id=enrollment.id,
            students=student_name,
            notes=f'自动排课 - 报名#{enrollment.id}',
            **delivery_fields,
        )
        sync_schedule_student_snapshot(schedule, enrollment=enrollment, preserve_history=False)
        db.session.add(schedule)
        db.session.flush()
        ensure_schedule_feedback_todo(schedule, created_by=enrollment.teacher_id)
        created_count += 1

    enrollment.status = 'confirmed'
    db.session.flush()
    sync_enrollment_status(enrollment)
    complete_replan_workflows_for_enrollment(enrollment.id)
    db.session.commit()
    return True, f'已生成 {created_count} 节课程', created_count


def send_leave_status_notification(leave_request):
    from modules.auth.models import ChatMessage

    enrollment = leave_request.enrollment or (leave_request.schedule.enrollment if leave_request.schedule else None)
    profile = enrollment.student_profile if enrollment else None
    if not profile or not profile.user_id:
        return

    status_text = '已批准' if leave_request.status == 'approved' else '已驳回'
    course_name = leave_request.schedule.course_name if leave_request.schedule else ''
    leave_date = leave_request.leave_date.isoformat() if leave_request.leave_date else ''
    content = f'[请假审批][{course_name}] 你在 {leave_date} 的请假申请{status_text}。'
    if leave_request.status == 'rejected' and leave_request.decision_comment:
        content += f' 处理说明：{leave_request.decision_comment}'
    db.session.add(ChatMessage(
        sender_id=leave_request.approved_by,
        receiver_id=profile.user_id,
        enrollment_id=enrollment.id if enrollment else None,
        content=content,
        is_read=False,
    ))


def process_leave_request_decision(leave_request, actor, *, approve, decision_comment=''):
    from modules.auth.workflow_services import cancel_schedule_feedback_todo, ensure_leave_makeup_workflow

    if not leave_request:
        return {
            'success': False,
            'status_code': 404,
            'error': '请假记录不存在',
        }
    if not user_can_approve_leave(actor, leave_request):
        return {
            'success': False,
            'status_code': 403,
            'error': '无权审批该请假申请',
        }
    if leave_request.status != 'pending':
        return {
            'success': False,
            'status_code': 400,
            'error': '该请假申请已处理',
        }

    cleaned_comment = (decision_comment or '').strip()
    if not approve and not cleaned_comment:
        return {
            'success': False,
            'status_code': 400,
            'error': '请先填写处理说明',
        }

    leave_request.status = 'approved' if approve else 'rejected'
    leave_request.decision_comment = cleaned_comment or None
    leave_request.approved_by = getattr(actor, 'id', None)
    send_leave_status_notification(leave_request)

    linked_enrollment = leave_request.enrollment or (
        leave_request.schedule.enrollment if leave_request.schedule else None
    )
    if linked_enrollment:
        sync_enrollment_status(linked_enrollment)

    workflow_todo = None
    if approve:
        if leave_request.schedule_id:
            cancel_schedule_feedback_todo(leave_request.schedule_id, reason='课程请假已批准')
        workflow_todo = ensure_leave_makeup_workflow(leave_request, actor_user=actor)

    db.session.commit()

    payload = build_leave_request_payload(leave_request, actor)
    if approve:
        payload['next_workflow_id'] = workflow_todo.id if workflow_todo else None
        payload['next_action_label'] = '提交补课建议'
        payload['next_action_hint'] = '请到待我提案里补充补课时间，教务会继续发送给学生确认。'

    return {
        'success': True,
        'status_code': 200,
        'data': payload,
    }


def save_course_feedback(schedule, teacher_id, data, *, submit=False):
    from modules.oa.models import CourseFeedback
    from modules.auth.models import Enrollment
    from modules.auth.workflow_services import complete_schedule_feedback_todo

    feedback = CourseFeedback.query.filter_by(
        schedule_id=schedule.id,
        teacher_id=teacher_id,
    ).first()
    if not feedback:
        feedback = CourseFeedback(schedule_id=schedule.id, teacher_id=teacher_id)
        db.session.add(feedback)

    feedback.summary = (data.get('summary') or '').strip() or None
    feedback.student_performance = (data.get('student_performance') or '').strip() or None
    feedback.homework = (data.get('homework') or '').strip() or None
    feedback.next_focus = (data.get('next_focus') or '').strip() or None

    if submit:
        if not feedback.summary:
            return False, '请先填写本次课程总结', None
        feedback.status = 'submitted'
        feedback.submitted_at = get_business_now()
    elif feedback.status != 'submitted':
        feedback.status = 'draft'

    db.session.flush()
    if submit:
        complete_schedule_feedback_todo(schedule.id)
        db.session.flush()
        enrollment = db.session.get(Enrollment, schedule.enrollment_id) if schedule.enrollment_id else None
        if enrollment:
            sync_enrollment_status(enrollment)
    db.session.commit()
    return True, '反馈已提交' if submit else '反馈草稿已保存', feedback


def export_enrollment_schedule_xlsx(enrollment_id):
    """导出排课课表为 .xlsx 文件。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from modules.auth.models import Enrollment
    from modules.oa.models import CourseSchedule

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return None, None, '报名记录不存在'

    if not enrollment.confirmed_slot:
        return None, None, '尚未确认排课方案'

    try:
        plan = normalize_plan(json.loads(enrollment.confirmed_slot), enrollment)
    except (json.JSONDecodeError, TypeError):
        return None, None, '排课数据异常'

    teacher_name = enrollment.teacher.display_name if enrollment.teacher else ''
    day_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    schedules = _linked_schedule_query(enrollment.id).order_by(CourseSchedule.date, CourseSchedule.time_start).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '课程表'

    header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    headers = ['序号', '日期', '星期', '开始时间', '结束时间', '课程名称', '授课教师', '学生']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    if schedules:
        for index, schedule in enumerate(schedules, 1):
            ws.append([
                index,
                schedule.date.isoformat(),
                day_names[schedule.day_of_week],
                schedule.time_start,
                schedule.time_end,
                schedule.course_name,
                schedule.teacher,
                schedule.students,
            ])
    else:
        for index, session in enumerate(plan.get('session_dates', []), 1):
            ws.append([
                index,
                session['date'],
                day_names[session['day_of_week']],
                session['time_start'],
                session['time_end'],
                enrollment.course_name,
                teacher_name,
                enrollment.student_name,
            ])

    for column in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'{enrollment.student_name}_{enrollment.course_name}_课程表.xlsx'
    return output, filename, None


def delete_student_user_hard(user_id):
    """仅允许硬删除学生账号，且需无强关联业务数据。"""
    from modules.auth.models import ChatMessage, Enrollment, LeaveRequest, StudentProfile, User
    from modules.oa.models import CourseSchedule

    user = db.session.get(User, user_id)
    if not user:
        return False, '用户不存在'
    if user.role != 'student':
        return False, '仅学生账号支持删除'

    profiles = StudentProfile.query.filter_by(user_id=user.id).all()
    profile_ids = [profile.id for profile in profiles]
    enrollment_ids = []

    if profile_ids:
        enrollment_ids = [
            row[0] for row in Enrollment.query.with_entities(Enrollment.id).filter(
                Enrollment.student_profile_id.in_(profile_ids)
            ).all()
        ]

    if enrollment_ids:
        return False, '该学生账号仍有关联的报名记录，请先删除相关报名'

    linked_schedule_ids = []
    if profile_ids:
        linked_schedule_ids = [
            row[0] for row in CourseSchedule.query.with_entities(CourseSchedule.id).filter(
                student_schedule_profile_clause(profile_ids, schedule_model=CourseSchedule)
            ).all()
        ]

    if linked_schedule_ids:
        return False, '该学生账号仍有关联的正式课表，请先删除相关报名'

    if linked_schedule_ids and LeaveRequest.query.filter(LeaveRequest.schedule_id.in_(linked_schedule_ids)).count() > 0:
        return False, '该学生账号仍有关联的请假记录，请先删除相关报名'

    ChatMessage.query.filter(
        or_(ChatMessage.sender_id == user.id, ChatMessage.receiver_id == user.id)
    ).delete(synchronize_session=False)

    for profile in profiles:
        db.session.delete(profile)
    db.session.delete(user)
    db.session.commit()
    return True, '学生账号已删除'


def delete_enrollment_hard(enrollment_id):
    """硬删除报名，并联动清理课表、请假、待办和站内消息。"""
    from modules.auth.models import ChatMessage, Enrollment, FeedbackShareLink, LeaveRequest
    from modules.oa.models import CourseFeedback, CourseSchedule, OATodo, ScheduleMeetingMaterial

    enrollment = db.session.get(Enrollment, enrollment_id)
    if not enrollment:
        return False, '报名记录不存在'

    linked_schedules = _linked_schedule_query(enrollment.id, include_cancelled=True).all()
    schedule_ids = [schedule.id for schedule in linked_schedules]

    OATodo.query.filter(OATodo.enrollment_id == enrollment.id).delete(synchronize_session=False)

    if schedule_ids:
        OATodo.query.filter(OATodo.schedule_id.in_(schedule_ids)).delete(synchronize_session=False)
        LeaveRequest.query.filter(LeaveRequest.schedule_id.in_(schedule_ids)).delete(synchronize_session=False)
        ScheduleMeetingMaterial.query.filter(
            ScheduleMeetingMaterial.schedule_id.in_(schedule_ids)
        ).delete(synchronize_session=False)
        CourseFeedback.query.filter(CourseFeedback.schedule_id.in_(schedule_ids)).delete(synchronize_session=False)
        CourseSchedule.query.filter(CourseSchedule.id.in_(schedule_ids)).delete(synchronize_session=False)

    ChatMessage.query.filter(ChatMessage.enrollment_id == enrollment.id).delete(synchronize_session=False)
    FeedbackShareLink.query.filter(FeedbackShareLink.enrollment_id == enrollment.id).delete(synchronize_session=False)

    profile = enrollment.student_profile
    should_delete_profile = False
    if profile:
        remaining = Enrollment.query.filter(
            Enrollment.student_profile_id == profile.id,
            Enrollment.id != enrollment.id,
        ).count()
        should_delete_profile = remaining == 0
    db.session.delete(enrollment)
    if profile and should_delete_profile:
        db.session.delete(profile)

    db.session.commit()
    return True, f'报名已删除，并清理 {len(schedule_ids)} 节关联课程'


def backfill_schedule_relationships():
    """回填历史自动排课记录的 teacher_id / enrollment_id。"""
    from modules.auth.models import Enrollment, User
    from modules.oa.models import CourseSchedule

    updated = 0
    schedules = CourseSchedule.query.filter(
        or_(
            CourseSchedule.enrollment_id == None,
            CourseSchedule.teacher_id == None,
            CourseSchedule.student_profile_id_snapshot == None,
        )
    ).all()
    for schedule in schedules:
        changed = False
        enrollment = None

        if schedule.enrollment_id:
            enrollment = db.session.get(Enrollment, schedule.enrollment_id)

        if not enrollment and schedule.notes:
            legacy_enrollment_id = _extract_legacy_enrollment_id(schedule.notes)
            if legacy_enrollment_id:
                enrollment = db.session.get(Enrollment, legacy_enrollment_id)
                if enrollment:
                    schedule.enrollment_id = enrollment.id
                    changed = True

        if not schedule.teacher_id:
            if enrollment:
                schedule.teacher_id = enrollment.teacher_id
                changed = True
            elif schedule.teacher:
                teacher = User.query.filter(
                    or_(User.display_name == schedule.teacher, User.username == schedule.teacher)
                ).first()
                if teacher:
                    schedule.teacher_id = teacher.id
                    changed = True

        if schedule.student_profile_id_snapshot is None:
            previous_snapshot = schedule.student_profile_id_snapshot
            sync_schedule_student_snapshot(schedule, enrollment=enrollment, preserve_history=True)
            if schedule.student_profile_id_snapshot != previous_snapshot:
                changed = True

        if changed:
            updated += 1

    if updated:
        db.session.commit()
    return updated


def seed_staff_accounts():
    """将现有硬编码员工初始化为用户账号。"""
    from modules.auth.models import User

    staff_accounts = [
        ('admin', '管理员', 'admin', 'admin'),
        ('liyu', '李宇', 'admin', 'scf123'),
        ('fanxiaodong', '范晓东', 'admin', 'scf123'),
        ('zhouxing', '周行', 'admin', 'scf123'),
        ('baoruimin', '包睿旻', 'teacher', 'scf123'),
        ('liyijun', '黎怡君', 'teacher', 'scf123'),
        ('zhangyu', '张渝', 'teacher', 'scf123'),
        ('chenguanru', '陈冠如', 'teacher', 'scf123'),
        ('wangyanlong', '王艳龙', 'teacher', 'scf123'),
        ('lulaoshi', '卢老师', 'teacher', 'scf123'),
        ('tianpeng', '田鹏', 'teacher', 'scf123'),
        ('chendonghao', '陈东豪', 'teacher', 'scf123'),
    ]

    created = 0
    for username, display_name, role, password in staff_accounts:
        if User.query.filter_by(username=username).first():
            continue
        user = User(username=username, display_name=display_name, role=role)
        user.set_password(password)
        db.session.add(user)
        created += 1

    if created > 0:
        db.session.commit()
    return created
