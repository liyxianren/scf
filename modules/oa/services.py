"""Excel 课表导入、语义归一化与 OA 课表辅助工具。"""
import os
import json
import re
from collections import Counter
from datetime import date, timedelta

from flask import current_app
from sqlalchemy import or_
from werkzeug.utils import secure_filename

# Monkey-patch openpyxl DataValidation to accept 'id' kwarg
# (compatibility fix for Excel files created with newer Office versions)
import openpyxl.worksheet.datavalidation as _dv
_original_dv_init = _dv.DataValidation.__init__


def _patched_dv_init(self, *args, **kwargs):
    kwargs.pop('id', None)
    _original_dv_init(self, *args, **kwargs)


_dv.DataValidation.__init__ = _patched_dv_init

import openpyxl


MONTH_MAP = {
    '12月': 12, '1月': 1, '2月': 2, '3月': 3, '4月': 4, '5月': 5,
    '6月': 6, '7月': 7, '8月': 8, '9月': 9, '10月': 10, '11月': 11
}

# Time pattern: "10:00-12:00" or "10:00～12:00" or "10:00 - 12:00"
TIME_PATTERN = re.compile(
    r'(\d{1,2}[：:]\d{2})\s*[-~—]\s*(\d{1,2}[：:]\d{2})'
)

COLOR_SEMANTICS = {
    'FFD9E1F4': ('online', 'blue'),
    'FFFEE796': ('offline', 'orange'),
    'FFFFE699': ('offline', 'orange'),
    'FF63FFD8': ('special', 'teal'),
}

COLOR_TAG_TO_DELIVERY_MODE = {
    'blue': 'online',
    'orange': 'offline',
    'teal': 'special',
}

COURSE_KEYWORDS = (
    '项目', '课程', '辅导', '竞赛', '训练', '面试', '工作坊', '技术课', '工程',
    'USACO', 'USABO', 'BBO', 'AMC', 'AIME', 'AI', '科研', '编程', '数学',
    '物理', '化学', '生物', '口语', '写作', '夏校',
)

TEACHER_ALIAS_MAP = {
    '田老师': '田鹏',
    '范老师': '范晓东',
    '李老师': '李宇',
    '黎老师': '黎怡君',
    '刘、范老师': '刘硕、范晓东',
}


def _normalize_todo_text(value):
    """Normalize text fields so duplicate detection ignores whitespace noise."""
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().casefold()


def _normalize_todo_date(value):
    if not value:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value).strip()


def build_todo_dedup_key(todo):
    """Build a stable business key for todo deduplication."""
    getter = getattr(todo, 'get', None)
    if callable(getter):
        get_value = lambda key, default=None: todo.get(key, default)
    else:
        get_value = lambda key, default=None: getattr(todo, key, default)

    return (
        _normalize_todo_text(get_value('title')),
        _normalize_todo_text(get_value('description')),
        _normalize_todo_text(get_value('responsible_person')),
        _normalize_todo_date(get_value('due_date')),
        int(get_value('priority', 2) or 2),
        _normalize_todo_text(get_value('notes')),
        get_value('schedule_id'),
        bool(get_value('is_completed', False)),
    )


def deduplicate_todo_payloads(todos):
    """
    Remove duplicate todo payloads before they are inserted into the database.
    Keeps the first occurrence for a given business key.
    """
    seen = set()
    deduplicated = []
    removed_count = 0

    for todo in todos:
        dedup_key = build_todo_dedup_key(todo)
        if dedup_key in seen:
            removed_count += 1
            continue
        seen.add(dedup_key)
        deduplicated.append(todo)

    return deduplicated, removed_count


def build_exact_todo_key(todo):
    """Build an exact field-level key for cleaning existing duplicated rows."""
    getter = getattr(todo, 'get', None)
    if callable(getter):
        get_value = lambda key, default=None: todo.get(key, default)
    else:
        get_value = lambda key, default=None: getattr(todo, key, default)

    return (
        get_value('title'),
        get_value('description'),
        get_value('responsible_person'),
        bool(get_value('is_completed', False)),
        get_value('due_date'),
        int(get_value('priority', 2) or 2),
        get_value('notes'),
        get_value('schedule_id'),
    )


def cleanup_existing_exact_duplicate_todos():
    """
    Delete exact duplicate todo rows from the database, keeping the earliest row.
    """
    from extensions import db
    from modules.oa.models import OATodo

    seen = {}
    removed_ids = []
    todos = OATodo.query.order_by(OATodo.id.asc()).all()

    for todo in todos:
        dedup_key = build_exact_todo_key(todo)
        if dedup_key in seen:
            removed_ids.append(todo.id)
            db.session.delete(todo)
            continue
        seen[dedup_key] = todo.id

    if removed_ids:
        db.session.commit()

    return {
        'total_before': len(todos),
        'removed_count': len(removed_ids),
        'kept_count': len(todos) - len(removed_ids),
        'removed_ids': removed_ids,
    }


def _normalize_schedule_text(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().casefold()


def build_schedule_import_key(schedule):
    getter = getattr(schedule, 'get', None)
    if callable(getter):
        get_value = lambda key, default=None: schedule.get(key, default)
    else:
        get_value = lambda key, default=None: getattr(schedule, key, default)

    schedule_date = get_value('date')
    date_key = schedule_date.isoformat() if hasattr(schedule_date, 'isoformat') else str(schedule_date)
    teacher_name, _ = normalize_teacher_name(get_value('teacher'))
    return (
        date_key,
        str(get_value('time_start') or '').strip(),
        str(get_value('time_end') or '').strip(),
        teacher_name,
    )


def deduplicate_schedule_payloads(schedules):
    seen = set()
    deduplicated = []
    removed_count = 0

    for schedule in schedules:
        dedup_key = build_schedule_import_key(schedule)
        if dedup_key in seen:
            removed_count += 1
            continue
        seen.add(dedup_key)
        deduplicated.append(schedule)

    return deduplicated, removed_count


def parse_excel_date_serial(serial_number):
    """Convert Excel date serial number to Python date."""
    base_date = date(1899, 12, 30)
    return base_date + timedelta(days=int(serial_number))


def normalize_time(t):
    """Normalize time string: replace Chinese colon, ensure HH:MM format."""
    t = t.replace('：', ':')
    parts = t.split(':')
    if len(parts) == 2:
        return f"{int(parts[0]):02d}:{parts[1]}"
    return t


def _normalize_teacher_token(value):
    if value is None:
        return ''
    normalized = str(value).strip()
    normalized = normalized.replace('\u3000', ' ')
    normalized = normalized.replace('，', '、').replace(',', '、')
    normalized = normalized.replace(';', '、').replace('；', '、')
    normalized = re.sub(r'\s+', '', normalized)
    normalized = normalized.strip('、')
    if normalized in {'—', '-', '无', '待定'}:
        return ''
    return normalized


def _strip_inline_annotation(value):
    if value is None:
        return ''
    return re.sub(r'[（(][^）)]*[）)]', '', str(value)).strip()


def _get_known_teacher_names():
    from modules.auth.models import User

    names = set(TEACHER_ALIAS_MAP.keys()) | set(TEACHER_ALIAS_MAP.values())
    try:
        users = User.query.filter(
            User.is_active == True,
            User.role.in_(['teacher', 'admin']),
        ).all()
        names.update(user.display_name for user in users if user.display_name)
        names.update(user.username for user in users if user.username)
    except Exception:
        pass
    return names


def _split_people_tokens(value):
    base = _strip_inline_annotation(value)
    return [token.strip() for token in re.split(r'[、，,]+', base) if token.strip()]


def _is_name_like_token(value):
    token = str(value or '').strip().strip('-—')
    if not token:
        return False
    if token.endswith('老师'):
        return True
    if re.fullmatch(r'[\u4e00-\u9fff]{2,4}', token):
        return True
    if re.fullmatch(r'[A-Za-z][A-Za-z .-]{0,20}', token):
        return True
    return False


def _looks_like_course_line(value):
    text = str(value or '').strip()
    if not text:
        return False
    lowered = text.casefold()
    return any(keyword.casefold() in lowered for keyword in COURSE_KEYWORDS)


def _looks_like_teacher_line(value):
    text = _normalize_teacher_token(_strip_inline_annotation(value))
    if not text:
        return False

    known_names = _get_known_teacher_names()
    if text in known_names:
        return True
    if text in TEACHER_ALIAS_MAP:
        return True
    if '老师' in text:
        return True

    tokens = _split_people_tokens(text)
    if not tokens:
        return False
    if all(token in known_names or token in TEACHER_ALIAS_MAP or token.endswith('老师') for token in tokens):
        return True
    return any(token in known_names for token in tokens)


def _looks_like_student_line(value):
    text = str(value or '').strip()
    if not text or _looks_like_course_line(text) or _looks_like_teacher_line(text):
        return False

    tokens = _split_people_tokens(text)
    if not tokens:
        return False
    if len(tokens) >= 2:
        return all(_is_name_like_token(token) for token in tokens)
    return _is_name_like_token(tokens[0])


def _extract_teacher_and_metadata(extra_lines, inline_teacher=''):
    teacher = inline_teacher.strip()
    remaining_lines = [line.strip() for line in extra_lines if str(line or '').strip()]

    if not teacher:
        for index, line in enumerate(remaining_lines):
            if _looks_like_teacher_line(line):
                teacher = line.strip()
                remaining_lines = remaining_lines[:index] + remaining_lines[index + 1:]
                break

    course_lines = []
    student_lines = []
    unknown_lines = []

    for line in remaining_lines:
        if _looks_like_course_line(line):
            course_lines.append(line)
        elif _looks_like_student_line(line):
            student_lines.append(line)
        else:
            unknown_lines.append(line)

    if not course_lines and unknown_lines:
        course_lines.append(unknown_lines.pop(0))
    if not student_lines and unknown_lines:
        student_lines.append(unknown_lines.pop(-1))

    course_name = ' / '.join(course_lines).strip()
    students = '、'.join(student_lines).strip()
    return teacher, course_name, students


def _extract_time_ranges(line):
    matches = list(TIME_PATTERN.finditer(line))
    if not matches:
        return [], ''

    time_ranges = [
        (normalize_time(match.group(1)), normalize_time(match.group(2)))
        for match in matches
    ]
    trailing_text = line[matches[-1].end():].strip()
    return time_ranges, trailing_text


def normalize_teacher_name(value):
    normalized = _normalize_teacher_token(value)
    base = _normalize_teacher_token(_strip_inline_annotation(normalized))
    known_names = _get_known_teacher_names()

    canonical = TEACHER_ALIAS_MAP.get(normalized, normalized)
    if canonical == normalized and base:
        canonical = TEACHER_ALIAS_MAP.get(base, base)

    if canonical.endswith('老师'):
        stripped = canonical[:-2]
        if stripped in known_names:
            canonical = stripped

    alias_hit = normalized if normalized and canonical != normalized else None
    return canonical, alias_hit


def resolve_schedule_teacher_user(teacher_name):
    from modules.auth.models import User

    canonical_name, _ = normalize_teacher_name(teacher_name)
    if not canonical_name:
        return None, '缺少授课教师'

    teacher_user = User.query.filter(
        User.is_active == True,
        User.role.in_(['teacher', 'admin']),
        or_(User.display_name == canonical_name, User.username == canonical_name),
    ).first()
    if not teacher_user:
        return None, f'未找到老师: {canonical_name}'
    return teacher_user, None


def resolve_schedule_teacher_reference(teacher_name):
    canonical_name, alias_hit = normalize_teacher_name(teacher_name)
    teacher_user, error = resolve_schedule_teacher_user(canonical_name)
    if error:
        return None, canonical_name, alias_hit, error
    return teacher_user, teacher_user.display_name, alias_hit, None


def delivery_mode_from_color_tag(color_tag):
    return COLOR_TAG_TO_DELIVERY_MODE.get((color_tag or '').strip().lower(), 'unknown')


def _time_to_minutes(time_str):
    if not time_str:
        return None
    try:
        hour, minute = str(time_str).split(':', 1)
        return int(hour) * 60 + int(minute)
    except (ValueError, TypeError):
        return None


def _time_ranges_overlap(start_a, end_a, start_b, end_b):
    a_start = _time_to_minutes(start_a)
    a_end = _time_to_minutes(end_a)
    b_start = _time_to_minutes(start_b)
    b_end = _time_to_minutes(end_b)
    if None in {a_start, a_end, b_start, b_end}:
        return False
    return max(a_start, b_start) < min(a_end, b_end)


def find_schedule_conflicts(
    *,
    course_date=None,
    time_start=None,
    time_end=None,
    teacher_id=None,
    teacher_name=None,
    enrollment_id=None,
    student_profile_id=None,
    exclude_schedule_id=None,
):
    from extensions import db
    from modules.auth.models import Enrollment
    from modules.oa.models import CourseSchedule

    result = {
        'error': None,
        'teacher': [],
        'enrollment': [],
        'student': [],
    }

    if not course_date or not time_start or not time_end or not (teacher_id or teacher_name):
        result['error'] = '缺少冲突校验所需的日期、时间或教师信息'
        result['all'] = []
        return result
    if _time_to_minutes(time_end) <= _time_to_minutes(time_start):
        result['error'] = '结束时间必须晚于开始时间'
        result['all'] = []
        return result

    teacher_clauses = []
    if teacher_id:
        teacher_clauses.append(CourseSchedule.teacher_id == teacher_id)
    if teacher_name:
        teacher_clauses.append(CourseSchedule.teacher == teacher_name)
    teacher_query = CourseSchedule.query.filter(
        CourseSchedule.date == course_date,
        or_(*teacher_clauses),
    )
    if exclude_schedule_id:
        teacher_query = teacher_query.filter(CourseSchedule.id != exclude_schedule_id)
    result['teacher'] = [
        schedule for schedule in teacher_query.all()
        if _time_ranges_overlap(time_start, time_end, schedule.time_start, schedule.time_end)
    ]

    if enrollment_id:
        enrollment_query = CourseSchedule.query.filter(
            CourseSchedule.enrollment_id == enrollment_id,
            CourseSchedule.date == course_date,
        )
        if exclude_schedule_id:
            enrollment_query = enrollment_query.filter(CourseSchedule.id != exclude_schedule_id)
        result['enrollment'] = [
            schedule for schedule in enrollment_query.all()
            if _time_ranges_overlap(time_start, time_end, schedule.time_start, schedule.time_end)
        ]

    if student_profile_id is None and enrollment_id:
        enrollment = db.session.get(Enrollment, enrollment_id)
        student_profile_id = enrollment.student_profile_id if enrollment else None

    if student_profile_id:
        student_query = CourseSchedule.query.join(
            Enrollment,
            CourseSchedule.enrollment_id == Enrollment.id,
        ).filter(
            Enrollment.student_profile_id == student_profile_id,
            CourseSchedule.date == course_date,
        )
        if enrollment_id:
            student_query = student_query.filter(Enrollment.id != enrollment_id)
        if exclude_schedule_id:
            student_query = student_query.filter(CourseSchedule.id != exclude_schedule_id)
        result['student'] = [
            schedule for schedule in student_query.all()
            if _time_ranges_overlap(time_start, time_end, schedule.time_start, schedule.time_end)
        ]

    deduped_conflicts = []
    seen_ids = set()
    for bucket in ('teacher', 'enrollment', 'student'):
        for schedule in result[bucket]:
            if schedule.id in seen_ids:
                continue
            seen_ids.add(schedule.id)
            deduped_conflicts.append(schedule)
    result['all'] = deduped_conflicts
    return result


def validate_schedule_conflicts(**kwargs):
    schedule_id = kwargs.pop('schedule_id', None)
    exclude_schedule_id = kwargs.pop('exclude_schedule_id', None)
    if exclude_schedule_id is None:
        exclude_schedule_id = schedule_id
    conflicts = find_schedule_conflicts(
        exclude_schedule_id=exclude_schedule_id,
        **kwargs,
    )
    if conflicts['error']:
        return conflicts['error']
    if conflicts['teacher']:
        return '该老师在相同时段已有课程安排'
    if conflicts['enrollment']:
        return '该报名在相同时段已有课程安排'
    if conflicts['student']:
        return '该学生在其他报名里同一时段已有课程安排'
    return None


def _resolve_unique_teacher_user(teacher_name):
    from modules.auth.models import User

    canonical_name, alias_hit = normalize_teacher_name(teacher_name)
    if not canonical_name:
        return None, canonical_name, alias_hit, 'missing'

    candidates = User.query.filter(
        User.is_active == True,
        User.role.in_(['teacher', 'admin']),
        or_(User.display_name == canonical_name, User.username == canonical_name),
    ).all()
    if len(candidates) == 1:
        return candidates[0], candidates[0].display_name, alias_hit, None
    if len(candidates) > 1:
        return None, canonical_name, alias_hit, 'ambiguous'
    return None, canonical_name, alias_hit, 'missing'


def _serialize_enrollment_candidate(enrollment):
    return {
        'id': enrollment.id,
        'student_name': enrollment.student_name,
        'course_name': enrollment.course_name,
        'teacher_id': enrollment.teacher_id,
        'teacher_name': enrollment.teacher.display_name if enrollment.teacher else None,
        'student_profile_id': enrollment.student_profile_id,
        'status': enrollment.status,
    }


def _find_import_enrollment_candidates(student_name, course_name):
    from modules.auth.models import Enrollment

    student_name = (student_name or '').strip()
    course_name = (course_name or '').strip()
    if not student_name or not course_name:
        return []

    return Enrollment.query.filter(
        Enrollment.student_name == student_name,
        Enrollment.course_name == course_name,
    ).order_by(Enrollment.id.asc()).all()


def _find_safe_import_enrollment(teacher_user, student_name, course_name):
    if not teacher_user:
        return None, []

    candidates = _find_import_enrollment_candidates(student_name, course_name)
    teacher_candidates = [enrollment for enrollment in candidates if enrollment.teacher_id == teacher_user.id]
    if len(teacher_candidates) == 1:
        return teacher_candidates[0], candidates
    return None, candidates


def _build_import_binding_todo_payload(schedule, *, issue_type, issue_message, candidate_enrollments):
    return {
        'schedule_id': schedule.id,
        'issue_type': issue_type,
        'issue_message': issue_message,
        'candidate_enrollments': candidate_enrollments,
        'schedule': {
            'id': schedule.id,
            'date': schedule.date.isoformat() if schedule.date else None,
            'time_start': schedule.time_start,
            'time_end': schedule.time_end,
            'teacher': schedule.teacher,
            'course_name': schedule.course_name,
            'students': schedule.students,
            'teacher_id': schedule.teacher_id,
            'enrollment_id': schedule.enrollment_id,
        },
    }


def _upsert_import_binding_todo(schedule, *, issue_type, issue_message, candidate_enrollments, created_by=None):
    from extensions import db
    from modules.oa.models import OATodo

    payload = _build_import_binding_todo_payload(
        schedule,
        issue_type=issue_type,
        issue_message=issue_message,
        candidate_enrollments=candidate_enrollments,
    )

    todo = OATodo.query.filter_by(
        schedule_id=schedule.id,
        todo_type=OATodo.TODO_TYPE_EXCEL_IMPORT,
    ).first()
    created = False
    if not todo:
        todo = OATodo(
            title=f'Excel 待绑定：{schedule.course_name}',
            description=f'{schedule.date.isoformat()} {schedule.time_start}-{schedule.time_end} 需要人工绑定报名',
            responsible_person='教务',
            due_date=schedule.date,
            priority=2,
            schedule_id=schedule.id,
            todo_type=OATodo.TODO_TYPE_EXCEL_IMPORT,
            created_by=created_by,
        )
        db.session.add(todo)
        db.session.flush()
        created = True

    todo.title = f'Excel 待绑定：{schedule.course_name}'
    todo.description = f'{schedule.date.isoformat()} {schedule.time_start}-{schedule.time_end} 需要人工绑定报名'
    todo.responsible_person = '教务'
    todo.due_date = schedule.date
    todo.priority = 2
    todo.schedule_id = schedule.id
    todo.todo_type = OATodo.TODO_TYPE_EXCEL_IMPORT
    todo.is_completed = False
    todo.notes = issue_message
    todo.set_payload_data(payload)
    return todo, created


def _complete_import_binding_todo(schedule, *, note=None):
    from extensions import db
    from modules.oa.models import OATodo

    todo = OATodo.query.filter_by(
        schedule_id=schedule.id,
        todo_type=OATodo.TODO_TYPE_EXCEL_IMPORT,
    ).first()
    if not todo:
        return None

    todo.is_completed = True
    if note:
        todo.notes = note
    db.session.add(todo)
    return todo


def _find_import_schedule_conflicts(
    course_date,
    time_start,
    time_end,
    *,
    teacher_id=None,
    teacher_name=None,
    enrollment_id=None,
    student_profile_id=None,
    exclude_schedule_id=None,
):
    return find_schedule_conflicts(
        course_date=course_date,
        time_start=time_start,
        time_end=time_end,
        teacher_id=teacher_id,
        teacher_name=teacher_name,
        enrollment_id=enrollment_id,
        student_profile_id=student_profile_id,
        exclude_schedule_id=exclude_schedule_id,
    )


def _preserve_import_row_artifacts(existing_schedule, *, touched_schedule_ids, touched_todo_ids):
    from modules.oa.models import OATodo

    if not existing_schedule:
        return

    touched_schedule_ids.add(existing_schedule.id)
    for todo in OATodo.query.filter_by(
        schedule_id=existing_schedule.id,
        todo_type=OATodo.TODO_TYPE_EXCEL_IMPORT,
    ).all():
        touched_todo_ids.add(todo.id)


def _extract_fill_rgb(cell):
    fill = getattr(cell, 'fill', None)
    if not fill or getattr(fill, 'patternType', None) in (None, 'none'):
        return None

    color = getattr(fill, 'fgColor', None)
    if not color:
        return None

    color_type = getattr(color, 'type', None)
    if color_type == 'rgb' and color.rgb:
        return str(color.rgb).upper()
    return None


def _resolve_color_semantics(fill_rgb):
    if not fill_rgb:
        return 'unknown', 'blue', None

    semantics = COLOR_SEMANTICS.get(fill_rgb.upper())
    if semantics:
        return semantics[0], semantics[1], None
    return 'unknown', 'blue', fill_rgb.upper()


def parse_course_cell(cell_value):
    """Parse a multi-line course cell into structured course entries."""
    if not cell_value or not isinstance(cell_value, str):
        return []

    cell_value = cell_value.strip()
    if not cell_value:
        return []

    lines = [l.strip() for l in cell_value.split('\n') if l.strip()]
    courses = []
    i = 0

    while i < len(lines):
        line = lines[i]
        time_ranges, inline_teacher = _extract_time_ranges(line)

        if time_ranges:
            j = i + 1
            extra_lines = []
            while j < len(lines):
                if TIME_PATTERN.search(lines[j]):
                    break
                extra_lines.append(lines[j])
                j += 1

            teacher, course_name, students = _extract_teacher_and_metadata(extra_lines, inline_teacher=inline_teacher)
            for time_start, time_end in time_ranges:
                courses.append({
                    'time_start': time_start,
                    'time_end': time_end,
                    'teacher': teacher,
                    'course_name': course_name,
                    'students': students,
                })

            i = j
        else:
            i += 1

    return courses


def _fix_date_year(d, expected_year):
    """Correct the year of a parsed date if it doesn't match the expected year."""
    if d.year == expected_year:
        return d
    try:
        return d.replace(year=expected_year)
    except ValueError:
        return date(expected_year, d.month, 28)


def _read_week_dates(row, expected_year):
    """Read 7 date values from the week separator row (columns A-G)."""
    week_dates = []
    for ci in range(7):
        if ci < len(row) and row[ci].value is not None:
            try:
                s = int(float(row[ci].value))
                if 40000 < s < 50000:
                    d = parse_excel_date_serial(s)
                    if expected_year:
                        d = _fix_date_year(d, expected_year)
                    week_dates.append(d)
                else:
                    week_dates.append(None)
            except (ValueError, TypeError):
                week_dates.append(None)
        else:
            week_dates.append(None)
    return week_dates


def import_schedule_from_excel(file_path, original_filename=None):
    """Parse the Excel schedule file into structured schedule and todo payloads."""
    from modules.oa.models import OATodo

    wb = openpyxl.load_workbook(file_path, data_only=True)
    schedules = []
    todos = []
    warnings = []
    teacher_alias_hits = Counter()

    filename = original_filename or os.path.basename(file_path)
    target_year = None
    year_match = re.search(r'(20\d{2})', filename)
    if year_match:
        target_year = int(year_match.group(1))

    for sheet_name in wb.sheetnames:
        if sheet_name not in MONTH_MAP:
            continue

        ws = wb[sheet_name]
        sheet_month = MONTH_MAP[sheet_name]
        expected_year = target_year
        if target_year and sheet_month == 12:
            expected_year = target_year - 1

        week_dates = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if len(row) < 7:
                continue

            cell_a_val = row[0].value
            is_week_separator = False

            if cell_a_val is not None:
                try:
                    serial = int(float(cell_a_val))
                    if 40000 < serial < 50000:
                        week_dates = _read_week_dates(row, expected_year)
                        is_week_separator = True
                except (ValueError, TypeError):
                    pass

            if is_week_separator or week_dates is None:
                continue

            for col_idx in range(7):
                if col_idx >= len(row) or col_idx >= len(week_dates):
                    break

                course_date = week_dates[col_idx]
                if course_date is None:
                    continue

                cell = row[col_idx]
                if cell is None or cell.value is None:
                    continue

                cell_value = str(cell.value).strip()
                if not cell_value:
                    continue

                parsed = parse_course_cell(cell_value)
                if not parsed:
                    continue

                fill_rgb = _extract_fill_rgb(cell)
                delivery_mode, color_tag, unknown_color = _resolve_color_semantics(fill_rgb)
                if unknown_color:
                    warnings.append(
                        f'{sheet_name}!{cell.coordinate} 颜色 {unknown_color} 未识别，已按 unknown 导入'
                    )

                for course in parsed:
                    teacher_name, alias_hit = normalize_teacher_name(course['teacher'])
                    if alias_hit:
                        teacher_alias_hits[f'{alias_hit}->{teacher_name}'] += 1

                    course_name = (course['course_name'] or '').strip()
                    if not course_name:
                        warnings.append(
                            f'{sheet_name}!{cell.coordinate} {course_date.isoformat()} {course["time_start"]}-{course["time_end"]} 缺少课程名称，已使用“未命名课程”'
                        )
                        course_name = '未命名课程'

                    if not teacher_name:
                        warnings.append(
                            f'{sheet_name}!{cell.coordinate} {course_date.isoformat()} {course["time_start"]}-{course["time_end"]} 缺少教师姓名，已标记为待匹配教师'
                        )
                        teacher_name = '待匹配教师'
                        teacher_user = None
                    else:
                        teacher_user, resolved_teacher_name, _, teacher_error = _resolve_unique_teacher_user(teacher_name)
                        if teacher_error:
                            warnings.append(
                                f'{sheet_name}!{cell.coordinate} 教师“{teacher_name}”无法唯一匹配，已按文本导入'
                            )
                            teacher_name = resolved_teacher_name or teacher_name
                            teacher_user = None
                        else:
                            teacher_name = resolved_teacher_name

                    schedules.append({
                        'date': course_date,
                        'day_of_week': course_date.weekday(),
                        'time_start': course['time_start'],
                        'time_end': course['time_end'],
                        'teacher': teacher_user.display_name if teacher_user else teacher_name,
                        'teacher_id': teacher_user.id if teacher_user else None,
                        'course_name': course_name,
                        'students': (course['students'] or '').strip(),
                        'color_tag': color_tag,
                        'delivery_mode': delivery_mode,
                    })

            if len(row) > 7:
                todo_cell = row[7]
                if todo_cell and todo_cell.value:
                    todo_text = str(todo_cell.value).strip()
                    if todo_text:
                        is_completed = False
                        if len(row) > 8 and row[8].value is not None:
                            try:
                                is_completed = bool(int(float(row[8].value)))
                            except (ValueError, TypeError):
                                pass

                        person = ''
                        if len(row) > 9 and row[9].value:
                            person = str(row[9].value).strip()

                        notes = ''
                        if len(row) > 10 and row[10].value:
                            notes = str(row[10].value).strip()

                        week_ref_date = next((d for d in week_dates if d), None)
                        todos.append({
                            'title': todo_text,
                            'is_completed': is_completed,
                            'responsible_person': person,
                            'notes': notes,
                            'due_date': week_ref_date,
                            'todo_type': OATodo.TODO_TYPE_EXCEL_IMPORT,
                        })

    wb.close()
    return {
        'schedules': schedules,
        'todos': todos,
        'warnings': warnings,
        'teacher_alias_hits': dict(teacher_alias_hits),
    }


def _resolve_data_root():
    try:
        db_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
    except RuntimeError:
        db_uri = ''
    if db_uri.startswith('sqlite:///'):
        db_path = db_uri.replace('sqlite:///', '', 1)
        db_path = os.path.abspath(db_path)
        directory = os.path.dirname(db_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
            return directory

    if os.path.isdir('/data'):
        return '/data'

    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))
    return os.path.join(project_root, 'data')


def get_schedule_import_root():
    root = os.path.join(_resolve_data_root(), 'imports', 'schedules')
    os.makedirs(root, exist_ok=True)
    return root


def save_schedule_import_file(file_storage, import_run_id, original_filename=None):
    if not file_storage:
        return None

    original_name = original_filename or getattr(file_storage, 'filename', '') or ''
    _, extension = os.path.splitext(original_name)
    extension = extension.lower() if extension else '.xlsx'
    filename = f'original{extension}'
    import_dir = os.path.join(get_schedule_import_root(), str(import_run_id))
    os.makedirs(import_dir, exist_ok=True)
    saved_path = os.path.join(import_dir, secure_filename(filename) or f'original{extension}')
    file_storage.save(saved_path)
    return saved_path


def _choose_preferred_schedule(existing, candidate):
    if existing is None:
        return candidate
    if existing.import_run_id is None and candidate.import_run_id is not None:
        return candidate
    return existing


def _apply_imported_schedule_payload(schedule, payload, *, import_run_id):
    schedule.date = payload['date']
    schedule.day_of_week = payload['day_of_week']
    schedule.time_start = payload['time_start']
    schedule.time_end = payload['time_end']
    schedule.teacher = payload['teacher']
    schedule.teacher_id = payload['teacher_id']
    schedule.course_name = payload['course_name']
    schedule.students = payload.get('students') or ''
    schedule.color_tag = payload.get('color_tag') or schedule.color_tag or 'blue'
    schedule.delivery_mode = payload.get('delivery_mode') or schedule.delivery_mode or 'unknown'
    schedule.import_run_id = import_run_id


def _apply_imported_todo_payload(todo, payload):
    todo.title = payload['title']
    todo.description = payload.get('description', todo.description)
    todo.responsible_person = payload.get('responsible_person', todo.responsible_person)
    todo.is_completed = bool(payload.get('is_completed', False))
    todo.due_date = payload.get('due_date')
    todo.priority = int(payload.get('priority', todo.priority or 2) or 2)
    todo.notes = payload.get('notes', todo.notes)
    todo.schedule_id = payload.get('schedule_id')
    todo.todo_type = payload.get('todo_type', todo.todo_type)
    if 'payload' in payload:
        todo.set_payload_data(payload.get('payload'))


def _can_delete_imported_schedule(schedule):
    return not schedule.feedback and not schedule.leave_requests and not schedule.todos


def _summarize_schedule(schedule):
    return f'{schedule.date.isoformat()} {schedule.time_start}-{schedule.time_end} {schedule.teacher} / {schedule.course_name}'


def apply_schedule_excel_import(file_storage, *, uploaded_by=None):
    from extensions import db
    from modules.auth.models import Enrollment
    from modules.auth.services import sync_enrollment_status
    from modules.auth.workflow_services import cancel_schedule_feedback_todo, ensure_schedule_feedback_todo
    from modules.oa.models import CourseSchedule, OATodo, ScheduleImportRun

    run = ScheduleImportRun(
        original_filename=(getattr(file_storage, 'filename', None) or 'schedule.xlsx'),
        uploaded_by=uploaded_by,
        status='pending',
    )
    db.session.add(run)
    db.session.commit()

    try:
        stored_path = save_schedule_import_file(file_storage, run.id, original_filename=run.original_filename)
        run = db.session.get(ScheduleImportRun, run.id)
        run.stored_path = stored_path
        db.session.commit()

        parsed = import_schedule_from_excel(stored_path, original_filename=run.original_filename)
        schedules, removed_schedule_duplicates = deduplicate_schedule_payloads(parsed['schedules'])
        todos, removed_todo_duplicates = deduplicate_todo_payloads(parsed['todos'])
        warnings = list(parsed.get('warnings') or [])
        teacher_alias_hits = dict(parsed.get('teacher_alias_hits') or {})

        existing_schedules = CourseSchedule.query.order_by(CourseSchedule.id.asc()).all()
        schedule_map = {}
        for schedule in existing_schedules:
            key = build_schedule_import_key(schedule)
            schedule_map[key] = _choose_preferred_schedule(schedule_map.get(key), schedule)

        touched_schedule_ids = set()
        schedules_created = 0
        schedules_updated = 0
        schedules_deleted = 0
        binding_todos_created = 0
        unmatched_schedules = []
        conflict_rows = []
        touched_todo_ids = set()

        for payload in schedules:
            key = build_schedule_import_key(payload)
            matched_schedule = schedule_map.get(key)
            existing = matched_schedule if matched_schedule and matched_schedule.import_run_id is not None else None
            date_value = payload.get('date')
            schedule_teacher_name = (payload.get('teacher') or '').strip()
            teacher_user = None
            resolved_teacher_name = schedule_teacher_name
            if payload.get('teacher_id'):
                teacher_user = db.session.get(
                    __import__('modules.auth.models', fromlist=['User']).User,
                    payload['teacher_id'],
                )
                if teacher_user:
                    resolved_teacher_name = teacher_user.display_name or teacher_user.username or schedule_teacher_name
            elif schedule_teacher_name and '待匹配' not in schedule_teacher_name:
                teacher_user, resolved_teacher_name, _, _ = _resolve_unique_teacher_user(schedule_teacher_name)
            teacher_name_for_conflict = resolved_teacher_name if resolved_teacher_name and '待匹配' not in resolved_teacher_name else None

            candidate_enrollments = _find_import_enrollment_candidates(
                payload.get('students'),
                payload.get('course_name'),
            )
            safe_enrollment = None
            if teacher_user:
                safe_enrollment, candidate_enrollments = _find_safe_import_enrollment(
                    teacher_user,
                    payload.get('students'),
                    payload.get('course_name'),
                )
            elif candidate_enrollments:
                safe_candidates = [
                    enrollment for enrollment in candidate_enrollments
                    if enrollment.teacher_id is not None and payload.get('teacher_id') == enrollment.teacher_id
                ]
                if len(safe_candidates) == 1:
                    safe_enrollment = safe_candidates[0]

            conflict_target_id = existing.id if existing else None
            conflicts = _find_import_schedule_conflicts(
                payload.get('date'),
                payload.get('time_start'),
                payload.get('time_end'),
                teacher_id=teacher_user.id if teacher_user else payload.get('teacher_id'),
                teacher_name=teacher_name_for_conflict,
                enrollment_id=safe_enrollment.id if safe_enrollment else None,
                student_profile_id=safe_enrollment.student_profile_id if safe_enrollment else None,
                exclude_schedule_id=conflict_target_id,
            )
            if conflicts['error']:
                _preserve_import_row_artifacts(
                    existing,
                    touched_schedule_ids=touched_schedule_ids,
                    touched_todo_ids=touched_todo_ids,
                )
                warnings.append(
                    f'导入跳过异常课次：{date_value.isoformat() if date_value else ""} '
                    f'{payload.get("time_start")}-{payload.get("time_end")} '
                    f'{schedule_teacher_name or "待匹配教师"} / {payload.get("course_name")}，'
                    f'{conflicts["error"]}'
                )
                continue
            if conflicts['all']:
                _preserve_import_row_artifacts(
                    existing,
                    touched_schedule_ids=touched_schedule_ids,
                    touched_todo_ids=touched_todo_ids,
                )
                conflict_rows.append({
                    'date': payload.get('date').isoformat() if payload.get('date') else None,
                    'time_start': payload.get('time_start'),
                    'time_end': payload.get('time_end'),
                    'teacher': schedule_teacher_name,
                    'course_name': payload.get('course_name'),
                    'students': payload.get('students') or '',
                    'conflicting_schedule_ids': [schedule.id for schedule in conflicts['all']],
                    'conflicting_schedules': [
                        {
                            'id': schedule.id,
                            'date': schedule.date.isoformat() if schedule.date else None,
                            'time_start': schedule.time_start,
                            'time_end': schedule.time_end,
                            'teacher': schedule.teacher,
                            'course_name': schedule.course_name,
                        }
                        for schedule in conflicts['all']
                    ],
                })
                teacher_label = teacher_name_for_conflict or schedule_teacher_name or '待匹配教师'
                warnings.append(
                    f'导入跳过冲突课次：{date_value.isoformat() if date_value else ""} '
                    f'{payload.get("time_start")}-{payload.get("time_end")} '
                    f'{teacher_label} / {payload.get("course_name")} 与现有课表冲突'
                )
                continue

            if existing:
                previous_enrollment_id = existing.enrollment_id
                _apply_imported_schedule_payload(existing, payload, import_run_id=run.id)
                if teacher_user:
                    existing.teacher = resolved_teacher_name
                    existing.teacher_id = teacher_user.id
                elif payload.get('teacher_id'):
                    existing.teacher_id = payload.get('teacher_id')
                else:
                    existing.teacher_id = None

                if safe_enrollment:
                    existing.enrollment_id = safe_enrollment.id
                    ensure_schedule_feedback_todo(
                        existing,
                        created_by=uploaded_by,
                    )
                    binding_todo = _complete_import_binding_todo(
                        existing,
                        note='已自动绑定报名，待绑定任务已关闭',
                    )
                    if binding_todo:
                        touched_todo_ids.add(binding_todo.id)
                    sync_enrollment_status(safe_enrollment)
                else:
                    existing.enrollment_id = None
                    cancel_schedule_feedback_todo(existing.id, reason='课次待重新绑定报名')
                    issue_message = (
                        f'未找到唯一报名绑定: {existing.date.isoformat()} {existing.time_start}-{existing.time_end} '
                        f'{existing.teacher} / {existing.course_name}'
                    )
                    todo, created = _upsert_import_binding_todo(
                        existing,
                        issue_type='unmatched_enrollment',
                        issue_message=issue_message,
                        candidate_enrollments=[_serialize_enrollment_candidate(enrollment) for enrollment in candidate_enrollments],
                        created_by=uploaded_by,
                    )
                    touched_todo_ids.add(todo.id)
                    if created:
                        binding_todos_created += 1
                    unmatched_schedules.append({
                        'schedule_id': existing.id,
                        'date': existing.date.isoformat() if existing.date else None,
                        'time_start': existing.time_start,
                        'time_end': existing.time_end,
                        'teacher': existing.teacher,
                        'course_name': existing.course_name,
                        'issue_type': 'unmatched_enrollment',
                        'issue_message': issue_message,
                        'candidate_enrollments': [_serialize_enrollment_candidate(enrollment) for enrollment in candidate_enrollments],
                    })

                if previous_enrollment_id and previous_enrollment_id != existing.enrollment_id:
                    previous_enrollment = db.session.get(Enrollment, previous_enrollment_id)
                    if previous_enrollment:
                        sync_enrollment_status(previous_enrollment)
                schedules_updated += 1
                touched_schedule_ids.add(existing.id)
            else:
                schedule = CourseSchedule(
                    **payload,
                    import_run_id=run.id,
                )
                if teacher_user:
                    schedule.teacher = resolved_teacher_name
                    schedule.teacher_id = teacher_user.id
                elif payload.get('teacher_id'):
                    schedule.teacher_id = payload.get('teacher_id')
                else:
                    schedule.teacher_id = None

                if safe_enrollment:
                    schedule.enrollment_id = safe_enrollment.id
                else:
                    schedule.enrollment_id = None

                db.session.add(schedule)
                db.session.flush()
                schedule_map[key] = schedule
                touched_schedule_ids.add(schedule.id)
                schedules_created += 1

                if safe_enrollment:
                    ensure_schedule_feedback_todo(
                        schedule,
                        created_by=uploaded_by,
                    )
                    binding_todo = _complete_import_binding_todo(
                        schedule,
                        note='已自动绑定报名，待绑定任务已关闭',
                    )
                    if binding_todo:
                        touched_todo_ids.add(binding_todo.id)
                    sync_enrollment_status(safe_enrollment)
                else:
                    issue_message = (
                        f'未找到唯一报名绑定: {schedule.date.isoformat()} {schedule.time_start}-{schedule.time_end} '
                        f'{schedule.teacher} / {schedule.course_name}'
                    )
                    todo, created = _upsert_import_binding_todo(
                        schedule,
                        issue_type='unmatched_enrollment',
                        issue_message=issue_message,
                        candidate_enrollments=[_serialize_enrollment_candidate(enrollment) for enrollment in candidate_enrollments],
                        created_by=uploaded_by,
                    )
                    touched_todo_ids.add(todo.id)
                    if created:
                        binding_todos_created += 1
                    unmatched_schedules.append({
                        'schedule_id': schedule.id,
                        'date': schedule.date.isoformat() if schedule.date else None,
                        'time_start': schedule.time_start,
                        'time_end': schedule.time_end,
                        'teacher': schedule.teacher,
                        'course_name': schedule.course_name,
                        'issue_type': 'unmatched_enrollment',
                        'issue_message': issue_message,
                        'candidate_enrollments': [_serialize_enrollment_candidate(enrollment) for enrollment in candidate_enrollments],
                    })

        stale_imported_schedules = CourseSchedule.query.filter(
            CourseSchedule.import_run_id.isnot(None),
            CourseSchedule.import_run_id != run.id,
        ).all()
        for schedule in stale_imported_schedules:
            if schedule.id in touched_schedule_ids:
                continue
            if not _can_delete_imported_schedule(schedule):
                warnings.append(
                    f'保留历史导入课表 #{schedule.id}：{_summarize_schedule(schedule)}，因为已有反馈/请假/待办关联'
                )
                continue
            db.session.delete(schedule)
            schedules_deleted += 1

        existing_excel_todos = OATodo.query.filter(
            OATodo.todo_type == OATodo.TODO_TYPE_EXCEL_IMPORT
        ).order_by(OATodo.id.asc()).all()
        todo_map = {}
        for todo in existing_excel_todos:
            key = build_todo_dedup_key(todo)
            todo_map.setdefault(key, todo)

        excel_todos_created = 0
        excel_todos_updated = 0
        excel_todos_deleted = 0

        for payload in todos:
            normalized_payload = dict(payload)
            normalized_payload['responsible_person'] = OATodo.normalize_responsible_people(
                normalized_payload.get('responsible_person', '')
            )
            normalized_payload['todo_type'] = OATodo.TODO_TYPE_EXCEL_IMPORT

            key = build_todo_dedup_key(normalized_payload)
            existing = todo_map.get(key)
            if existing:
                _apply_imported_todo_payload(existing, normalized_payload)
                touched_todo_ids.add(existing.id)
                excel_todos_updated += 1
            else:
                todo = OATodo(**normalized_payload)
                db.session.add(todo)
                db.session.flush()
                todo_map[key] = todo
                touched_todo_ids.add(todo.id)
                excel_todos_created += 1

        for todo in existing_excel_todos:
            if todo.id in touched_todo_ids:
                continue
            db.session.delete(todo)
            excel_todos_deleted += 1

        for enrollment in Enrollment.query.all():
            sync_enrollment_status(enrollment)

        summary = {
            'import_id': run.id,
            'schedules_created': schedules_created,
            'schedules_updated': schedules_updated,
            'schedules_deleted': schedules_deleted,
            'excel_todos_created': excel_todos_created,
            'excel_todos_updated': excel_todos_updated,
            'excel_todos_deleted': excel_todos_deleted,
            'binding_todos_created': binding_todos_created,
            'unmatched_schedules': unmatched_schedules,
            'conflict_rows': conflict_rows,
            'removed_schedule_duplicates': removed_schedule_duplicates,
            'removed_todo_duplicates': removed_todo_duplicates,
            'warnings': warnings,
            'teacher_alias_hits': teacher_alias_hits,
            'schedules_count': schedules_created + schedules_updated,
            'todos_count': excel_todos_created + excel_todos_updated,
            'stored_path': stored_path,
        }

        run = db.session.get(ScheduleImportRun, run.id)
        run.status = 'completed'
        run.set_summary_data(summary)
        db.session.commit()
        return run, summary
    except Exception as exc:
        db.session.rollback()
        run = db.session.get(ScheduleImportRun, run.id)
        if run:
            run.status = 'failed'
            run.set_summary_data({'error': str(exc)})
            db.session.commit()
        raise


def backfill_schedule_semantics():
    from extensions import db
    from modules.oa.models import CourseSchedule

    schedules = CourseSchedule.query.order_by(CourseSchedule.id.asc()).all()
    updated = 0
    teacher_alias_fixed = 0
    teacher_ids_linked = 0
    delivery_modes_backfilled = 0

    for schedule in schedules:
        changed = False
        canonical_teacher, alias_hit = normalize_teacher_name(schedule.teacher)
        if canonical_teacher and canonical_teacher != (schedule.teacher or ''):
            schedule.teacher = canonical_teacher
            changed = True
            if alias_hit:
                teacher_alias_fixed += 1

        if schedule.teacher:
            teacher_user, error = resolve_schedule_teacher_user(schedule.teacher)
            if not error and teacher_user and schedule.teacher_id != teacher_user.id:
                schedule.teacher = teacher_user.display_name
                schedule.teacher_id = teacher_user.id
                changed = True
                teacher_ids_linked += 1

        delivery_mode = delivery_mode_from_color_tag(schedule.color_tag)
        if (schedule.delivery_mode or 'unknown') != delivery_mode:
            schedule.delivery_mode = delivery_mode
            changed = True
            delivery_modes_backfilled += 1

        if changed:
            updated += 1

    if updated:
        db.session.commit()

    return {
        'updated': updated,
        'teacher_alias_fixed': teacher_alias_fixed,
        'teacher_ids_linked': teacher_ids_linked,
        'delivery_modes_backfilled': delivery_modes_backfilled,
    }
